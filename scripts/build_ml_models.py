"""
build_ml_models.py
Diagnostic classification and regression models for the SSC/C-OPN dataset.

Tasks
-----
Classification
  1. PD vs Healthy Control (binary)
  2. PD vs AP  (binary, class-weighted for imbalance)
  3. AP subtype (PSP / MSA / DLB-other grouped, small-N caveat)

Regression
  4. UPDRS Part III (motor exam) from non-motor features
  5. PDQ-39 total from clinical features

Output: output/ml_results.xlsx  +  output/ml_plots/*.png

Run: python3 scripts/build_ml_models.py
"""

import os, re, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.pipeline          import Pipeline
from sklearn.preprocessing     import StandardScaler, LabelEncoder
from sklearn.impute            import SimpleImputer
from sklearn.linear_model      import LogisticRegression
from sklearn.ensemble          import RandomForestClassifier, RandomForestRegressor
from sklearn.model_selection   import StratifiedKFold, KFold, cross_val_predict
from sklearn.metrics           import (
    roc_auc_score, balanced_accuracy_score, classification_report,
    confusion_matrix, r2_score, mean_absolute_error
)
from xgboost import XGBClassifier, XGBRegressor

warnings.filterwarnings('ignore')

ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR  = os.path.join(ROOT, 'data')
OUT_DIR   = os.path.join(ROOT, 'output')
PLOT_DIR  = os.path.join(OUT_DIR, 'ml_plots')
os.makedirs(PLOT_DIR, exist_ok=True)

DIAG_LABEL = {0: 'PD', 1: 'PSP', 2: 'MSA', 3: 'CBS', 4: 'DLB', 6: 'ET', 7: 'RBD'}


# ── Helpers ───────────────────────────────────────────────────────────────────
def load(name):
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    return df

def first_col(df, *keywords):
    """Return first column whose name contains all keywords (case-insensitive)."""
    kws = [k.lower() for k in keywords]
    for c in df.columns:
        cl = c.lower()
        if all(k in cl for k in kws):
            return c
    return None

def save_fig(name):
    path = os.path.join(PLOT_DIR, name)
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f'    saved {path}')


# ── Load raw forms ─────────────────────────────────────────────────────────────
print('Loading data...')
enroll = load('Enrollement')
clin   = load('Clinical')
updrs  = load('MDS-UPDRS')
moca   = load('MoCA')
demo   = load('Demographic')
epi    = load('Epidemiological')
pdq    = load('PDQ 39')
scopa  = load('SCOPA')
neuro  = load('Neuropsychological')


# ── Build master participant table ─────────────────────────────────────────────
print('Building participant table...')

enrol_col  = first_col(enroll, 'nrolment', 'roup')
status_col = first_col(enroll, 'Study Status')
diag_col   = first_col(clin,   'Determined diagnosis')
pd_yn_col  = first_col(clin,   'Was the patient diagnosed')

# Active enrolled only
active = (
    enroll[enroll[status_col].str.contains('Enrolled', na=False)]
    [['Project key', enrol_col]].drop_duplicates('Project key')
    .rename(columns={enrol_col: 'enrol_group'})
)

def short_enrol(v):
    if pd.isna(v): return None
    if 'Healthy' in v: return 'HC'
    if 'Atypical' in v: return 'AP'
    if "Parkinson's Disease" in v or 'Maladie' in v: return 'PD'
    return v

active['group'] = active['enrol_group'].apply(short_enrol)

# Merge determined dx
clin_dx = clin[['Project key', diag_col]].drop_duplicates('Project key').copy()
clin_dx['det_code'] = pd.to_numeric(clin_dx[diag_col], errors='coerce')
clin_dx['Dx']       = clin_dx['det_code'].map(DIAG_LABEL)

base = active.merge(clin_dx[['Project key', 'det_code', 'Dx']], on='Project key', how='left')
print(f'  Active enrolled: {len(base):,}  |  With determined Dx: {base["Dx"].notna().sum():,}')


# ── Feature extraction ─────────────────────────────────────────────────────────
print('Extracting features...')

# -- MoCA subscores (7 subscores + total)
moca_score_cols = {
    'moca_visuospatial': first_col(moca, 'visuospatial'),
    'moca_naming':       first_col(moca, 'naming', 'score'),
    'moca_attention':    first_col(moca, 'attention', 'score'),
    'moca_language':     first_col(moca, 'language', 'score'),
    'moca_abstraction':  first_col(moca, 'abstraction', 'score'),
    'moca_recall':       first_col(moca, 'recall', 'score'),
    'moca_orientation':  first_col(moca, 'orientation', 'score'),
    'moca_total':        first_col(moca, 'total score'),
}
moca_feat = moca[['Project key'] + [v for v in moca_score_cols.values() if v]].copy()
moca_feat = moca_feat.drop_duplicates('Project key')
moca_feat = moca_feat.rename(columns={v: k for k, v in moca_score_cols.items() if v})
for k in moca_score_cols:
    if k in moca_feat.columns:
        moca_feat[k] = pd.to_numeric(moca_feat[k], errors='coerce')

# -- UPDRS Part totals (I–IV)
updrs_part_cols = {
    'updrs_part1': first_col(updrs, 'part i'),
    'updrs_part2': first_col(updrs, 'part ii'),
    'updrs_part3': first_col(updrs, 'part iii'),
    'updrs_part4': first_col(updrs, 'part iv'),
}
updrs_feat = updrs[['Project key'] + [v for v in updrs_part_cols.values() if v]].copy()
updrs_feat = updrs_feat.drop_duplicates('Project key')
updrs_feat = updrs_feat.rename(columns={v: k for k, v in updrs_part_cols.items() if v})
for k in updrs_part_cols:
    if k in updrs_feat.columns:
        updrs_feat[k] = pd.to_numeric(updrs_feat[k], errors='coerce')

# -- Demographics: age, sex, education years
age_col  = first_col(demo, 'age at study visit', 'automatic')
age_col  = age_col or first_col(demo, 'age at study visit')
sex_col  = first_col(demo, 'gender')
edu_col  = first_col(demo, 'years of education')
demo_feat = demo[['Project key'] + [c for c in [age_col, sex_col, edu_col] if c]].copy()
demo_feat = demo_feat.drop_duplicates('Project key')
demo_feat = demo_feat.rename(columns={
    age_col: 'age', sex_col: 'sex', edu_col: 'edu_years'
})
demo_feat['age']      = pd.to_numeric(demo_feat.get('age'),      errors='coerce')
demo_feat['edu_years']= pd.to_numeric(demo_feat.get('edu_years'),errors='coerce')
if 'sex' in demo_feat.columns:
    demo_feat['sex'] = demo_feat['sex'].astype(str).str.lower().str.strip()
    demo_feat['sex_binary'] = demo_feat['sex'].map(
        lambda v: 1 if 'male' in v and 'female' not in v else (0 if 'female' in v else np.nan)
    )
    demo_feat = demo_feat.drop(columns=['sex'])

# -- Clinical: disease duration, age at onset
dur_col   = first_col(clin, 'duration of disease at the time the clinical questionnaire')
onset_col = first_col(clin, 'age at onset')
onset_col = onset_col or first_col(clin, 'age at diagnosis')
clin_feat = clin[['Project key'] + [c for c in [dur_col, onset_col] if c]].copy()
clin_feat = clin_feat.drop_duplicates('Project key')
clin_feat = clin_feat.rename(columns={
    dur_col: 'disease_duration', onset_col: 'age_at_onset'
})
for c in ['disease_duration', 'age_at_onset']:
    if c in clin_feat.columns:
        clin_feat[c] = pd.to_numeric(clin_feat[c], errors='coerce')

# -- PDQ-39 total
pdq_total_col = first_col(pdq, 'total') or first_col(pdq, 'pdq')
if pdq_total_col:
    pdq_feat = pdq[['Project key', pdq_total_col]].drop_duplicates('Project key').copy()
    pdq_feat = pdq_feat.rename(columns={pdq_total_col: 'pdq39_total'})
    pdq_feat['pdq39_total'] = pd.to_numeric(pdq_feat['pdq39_total'], errors='coerce')
else:
    pdq_feat = pd.DataFrame(columns=['Project key', 'pdq39_total'])

# Merge all features into master
feat = base.copy()
for fdf in [moca_feat, updrs_feat, demo_feat, clin_feat, pdq_feat]:
    feat = feat.merge(fdf, on='Project key', how='left')

# All numeric feature columns available
ALL_FEATURES = [
    'moca_visuospatial', 'moca_naming', 'moca_attention', 'moca_language',
    'moca_abstraction', 'moca_recall', 'moca_orientation', 'moca_total',
    'updrs_part1', 'updrs_part2', 'updrs_part3', 'updrs_part4',
    'age', 'sex_binary', 'edu_years', 'disease_duration', 'age_at_onset',
]
ALL_FEATURES = [f for f in ALL_FEATURES if f in feat.columns]

# Non-motor features (for UPDRS Part III regression — exclude motor UPDRS parts)
NON_MOTOR = [
    'moca_visuospatial', 'moca_naming', 'moca_attention', 'moca_language',
    'moca_abstraction', 'moca_recall', 'moca_orientation', 'moca_total',
    'age', 'sex_binary', 'edu_years', 'disease_duration', 'age_at_onset',
    'updrs_part1',
]
NON_MOTOR = [f for f in NON_MOTOR if f in feat.columns]

print(f'  Feature columns available: {ALL_FEATURES}')


# ── Model pipeline factory ─────────────────────────────────────────────────────
def make_clf_pipelines(class_weight=None):
    return {
        'Logistic Regression': Pipeline([
            ('imp',   SimpleImputer(strategy='median')),
            ('scl',   StandardScaler()),
            ('clf',   LogisticRegression(max_iter=1000, class_weight=class_weight,
                                          random_state=42)),
        ]),
        'Random Forest': Pipeline([
            ('imp',   SimpleImputer(strategy='median')),
            ('clf',   RandomForestClassifier(n_estimators=300, class_weight=class_weight,
                                              random_state=42, n_jobs=-1)),
        ]),
        'XGBoost': Pipeline([
            ('imp',   SimpleImputer(strategy='median')),
            ('clf',   XGBClassifier(n_estimators=300, learning_rate=0.05,
                                     use_label_encoder=False, eval_metric='logloss',
                                     random_state=42, verbosity=0)),
        ]),
    }

def make_reg_pipelines():
    return {
        'Logistic Regression (Ridge)': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('scl', StandardScaler()),
            ('reg', __import__('sklearn.linear_model', fromlist=['Ridge']).Ridge()),
        ]),
        'Random Forest': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('reg', RandomForestRegressor(n_estimators=300, random_state=42, n_jobs=-1)),
        ]),
        'XGBoost': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('reg', XGBRegressor(n_estimators=300, learning_rate=0.05,
                                  random_state=42, verbosity=0)),
        ]),
    }


# ── Classification runner ──────────────────────────────────────────────────────
clf_results = []

def run_classification(task_name, df_task, label_col, features, n_splits=5,
                        class_weight=None, multiclass=False):
    print(f'\n  Task: {task_name}')
    df_task = df_task[features + [label_col]].dropna(subset=[label_col]).copy()

    le = LabelEncoder()
    y  = le.fit_transform(df_task[label_col])
    X  = df_task[features].values
    classes = le.classes_

    print(f'    N={len(y)}  classes={dict(zip(classes, np.bincount(y)))}')
    if len(np.unique(y)) < 2:
        print('    Skipping — only one class.')
        return

    cv = StratifiedKFold(n_splits=min(n_splits, np.bincount(y).min()), shuffle=True, random_state=42)
    pipelines = make_clf_pipelines(class_weight=class_weight)

    for model_name, pipe in pipelines.items():
        try:
            if multiclass:
                y_prob = cross_val_predict(pipe, X, y, cv=cv, method='predict_proba')
                y_pred = y_prob.argmax(axis=1)
                auc = roc_auc_score(y, y_prob, multi_class='ovr', average='macro')
            else:
                y_prob = cross_val_predict(pipe, X, y, cv=cv, method='predict_proba')[:, 1]
                y_pred = (y_prob >= 0.5).astype(int)
                auc = roc_auc_score(y, y_prob)

            bal_acc = balanced_accuracy_score(y, y_pred)
            row = {
                'Task':             task_name,
                'Model':            model_name,
                'N':                len(y),
                'Classes':          ' / '.join(str(c) for c in classes),
                'ROC-AUC':          round(auc, 3),
                'Balanced Acc':     round(bal_acc, 3),
            }
            # Per-class precision/recall from classification report
            rpt = classification_report(y, y_pred, target_names=classes, output_dict=True, zero_division=0)
            for cls in classes:
                row[f'F1 {cls}'] = round(rpt[cls]['f1-score'], 3)

            clf_results.append(row)
            print(f'    {model_name:30s}  AUC={auc:.3f}  BalAcc={bal_acc:.3f}')

            # Confusion matrix (last model only, best visual)
            if model_name == 'Random Forest':
                cm = confusion_matrix(y, y_pred)
                fig, ax = plt.subplots(figsize=(max(4, len(classes)), max(3, len(classes))))
                sns.heatmap(cm, annot=True, fmt='d', xticklabels=classes,
                            yticklabels=classes, cmap='Blues', ax=ax)
                ax.set_title(f'{task_name}\n{model_name} — Confusion Matrix (CV)')
                ax.set_xlabel('Predicted'); ax.set_ylabel('Actual')
                plt.tight_layout()
                fname = re.sub(r'[^a-zA-Z0-9]', '_', f'{task_name}_{model_name}') + '_cm.png'
                save_fig(fname)

        except Exception as e:
            print(f'    {model_name}: ERROR — {e}')

    # Feature importance from Random Forest (fit on full data for display)
    try:
        rf_pipe = make_clf_pipelines(class_weight=class_weight)['Random Forest']
        df_full = df_task[features + [label_col]].copy()
        X_full  = SimpleImputer(strategy='median').fit_transform(df_full[features].values)
        rf_pipe['clf'].fit(X_full, y)
        importances = rf_pipe['clf'].feature_importances_
        fi_df = pd.Series(importances, index=features).sort_values(ascending=True)
        fig, ax = plt.subplots(figsize=(6, max(3, len(features) * 0.35)))
        fi_df.plot.barh(ax=ax, color='steelblue')
        ax.set_title(f'{task_name}\nRandom Forest Feature Importance')
        ax.set_xlabel('Importance')
        plt.tight_layout()
        fname = re.sub(r'[^a-zA-Z0-9]', '_', task_name) + '_feature_importance.png'
        save_fig(fname)
    except Exception as e:
        print(f'    Feature importance: {e}')


# ── Regression runner ──────────────────────────────────────────────────────────
reg_results = []

def run_regression(task_name, df_task, target_col, features, n_splits=5):
    print(f'\n  Task: {task_name}')
    df_task = df_task[features + [target_col]].dropna(subset=[target_col]).copy()
    df_task = df_task.dropna(subset=features, thresh=len(features) // 2)

    y = df_task[target_col].values.astype(float)
    X = df_task[features].values
    print(f'    N={len(y)}  target mean={y.mean():.1f} ± {y.std():.1f}')

    if len(y) < 30:
        print('    Skipping — too few rows.')
        return

    cv = KFold(n_splits=min(n_splits, len(y)), shuffle=True, random_state=42)
    pipelines = make_reg_pipelines()

    for model_name, pipe in pipelines.items():
        try:
            y_pred = cross_val_predict(pipe, X, y, cv=cv)
            r2  = r2_score(y, y_pred)
            mae = mean_absolute_error(y, y_pred)
            reg_results.append({
                'Task':    task_name,
                'Model':   model_name,
                'N':       len(y),
                'R²':      round(r2, 3),
                'MAE':     round(mae, 2),
                'Target mean': round(float(y.mean()), 2),
                'Target std':  round(float(y.std()), 2),
            })
            print(f'    {model_name:35s}  R²={r2:.3f}  MAE={mae:.2f}')
        except Exception as e:
            print(f'    {model_name}: ERROR — {e}')

    # Scatter: actual vs predicted (Random Forest)
    try:
        rf_pipe = make_reg_pipelines()['Random Forest']
        y_pred_rf = cross_val_predict(rf_pipe, X, y, cv=cv)
        fig, ax = plt.subplots(figsize=(5, 4))
        ax.scatter(y, y_pred_rf, alpha=0.3, s=15, color='steelblue')
        lims = [min(y.min(), y_pred_rf.min()) - 1, max(y.max(), y_pred_rf.max()) + 1]
        ax.plot(lims, lims, 'r--', lw=1)
        ax.set_xlabel('Actual'); ax.set_ylabel('Predicted (CV)')
        ax.set_title(f'{task_name}\nRandom Forest — Actual vs Predicted')
        plt.tight_layout()
        fname = re.sub(r'[^a-zA-Z0-9]', '_', task_name) + '_actual_vs_pred.png'
        save_fig(fname)
    except Exception as e:
        print(f'    Scatter plot: {e}')


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 1 — PD vs Healthy Control
# ═══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*60)
print('CLASSIFICATION TASKS')
print('='*60)

task1 = feat[feat['group'].isin(['PD', 'HC'])].copy()
task1['label'] = task1['group']
run_classification(
    'Task 1 — PD vs Healthy Control',
    task1, 'label', ALL_FEATURES,
    class_weight='balanced',
)


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 2 — PD vs AP (all AP subtypes combined)
# ═══════════════════════════════════════════════════════════════════════════════
task2 = feat[feat['Dx'].notna() & feat['group'].isin(['PD', 'AP'])].copy()
task2['label'] = task2['group']
run_classification(
    'Task 2 — PD vs AP (all subtypes)',
    task2, 'label', ALL_FEATURES,
    class_weight='balanced',
)


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 3 — AP subtype classification (PSP / MSA / DLB+other)
# ═══════════════════════════════════════════════════════════════════════════════
ap_only = feat[feat['group'] == 'AP'].copy()
ap_only['ap_subtype'] = ap_only['Dx'].apply(
    lambda d: d if d in ('PSP', 'MSA') else ('DLB+other' if pd.notna(d) else np.nan)
)
run_classification(
    'Task 3 — AP subtypes (PSP / MSA / DLB+other)',
    ap_only, 'ap_subtype', ALL_FEATURES,
    n_splits=3,    # small N — use 3-fold
    multiclass=True,
)


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 4 — Predict UPDRS Part III (motor) from non-motor features
# ═══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*60)
print('REGRESSION TASKS')
print('='*60)

task4 = feat[feat['group'] == 'PD'].copy()   # PD patients only
run_regression(
    'Task 4 — UPDRS Part III from non-motor (PD only)',
    task4, 'updrs_part3', NON_MOTOR,
)


# ═══════════════════════════════════════════════════════════════════════════════
# TASK 5 — Predict PDQ-39 quality of life from clinical features
# ═══════════════════════════════════════════════════════════════════════════════
task5 = feat[feat['group'] == 'PD'].copy()
run_regression(
    'Task 5 — PDQ-39 from clinical features (PD only)',
    task5, 'pdq39_total', ALL_FEATURES,
)


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT — Excel results
# ═══════════════════════════════════════════════════════════════════════════════
print('\nWriting results...')
out_path = os.path.join(OUT_DIR, 'ml_results.xlsx')

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils  import get_column_letter

def style_results(ws):
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_w = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 40)

clf_df = pd.DataFrame(clf_results)
reg_df = pd.DataFrame(reg_results)

notes = pd.DataFrame([
    ['Task 1', 'PD vs HC', 'Binary', 'All features', 'Balanced class weight'],
    ['Task 2', 'PD vs AP (all subtypes)', 'Binary, severely imbalanced ~26:1',
     'All features', 'Balanced class weight — interpret with caution'],
    ['Task 3', 'AP subtype (PSP/MSA/DLB+other)', 'Multiclass, very small N (~56)',
     'All features', '3-fold CV, treat results as exploratory only'],
    ['Task 4', 'UPDRS Part III regression (PD)', 'Regression', 'Non-motor features only',
     'PD patients with UPDRS data'],
    ['Task 5', 'PDQ-39 regression (PD)', 'Regression', 'All features',
     'PD patients with PDQ-39 data (~468)'],
], columns=['Task', 'Description', 'Notes', 'Features', 'Caveats'])

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    clf_df.to_excel(writer, sheet_name='Classification Results', index=False)
    reg_df.to_excel(writer, sheet_name='Regression Results',     index=False)
    notes.to_excel(writer,  sheet_name='Task Notes',             index=False)

import openpyxl
wb = openpyxl.load_workbook(out_path)
for sn in wb.sheetnames:
    style_results(wb[sn])
wb.save(out_path)

print(f'\nDone.')
print(f'  Results → {out_path}')
print(f'  Plots   → {PLOT_DIR}/')
if clf_results:
    print('\nClassification summary:')
    print(clf_df[['Task', 'Model', 'N', 'ROC-AUC', 'Balanced Acc']].to_string(index=False))
if reg_results:
    print('\nRegression summary:')
    print(reg_df[['Task', 'Model', 'N', 'R²', 'MAE']].to_string(index=False))
