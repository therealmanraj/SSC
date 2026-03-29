"""
build_ml_v2.py
Full diagnostic classification pipeline using the Clinical (Labelled) sheet.

Label map: 0=PD  1=PSP  2=MSA  3=CBS  4=DLB  6=ET  7=RBD  10=HC

Tasks
-----
  A — PD vs HC                      (binary, balanced)
  B — PD vs AP  all subtypes        (binary, imbalanced ~17:1)
  C — PD vs HC vs AP                (3-class)
  D — AP subtype: PSP/MSA/DLB+other (multiclass, small-N, exploratory)

Features
--------
  Clinical  : disease duration, age at onset, BMI, 8 binary symptom flags
  MoCA      : 7 subscores + total
  MDS-UPDRS : Part I–IV totals
  SCOPA     : 6 autonomic subscores
  Demographic: age, sex, education years
  Neuro     : 8 key neuropsychological scores

Output
------
  output/ml_results_v2.xlsx   — metrics per task × model
  output/ml_plots_v2/         — confusion matrices, feature importance, SHAP

Run: python3 scripts/build_ml_v2.py
"""

import os, re, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import shap

from sklearn.pipeline        import Pipeline
from sklearn.preprocessing   import StandardScaler, LabelEncoder
from sklearn.impute          import SimpleImputer
from sklearn.linear_model    import LogisticRegression
from sklearn.ensemble        import RandomForestClassifier
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics         import (
    roc_auc_score, balanced_accuracy_score,
    classification_report, confusion_matrix,
)
from xgboost import XGBClassifier

warnings.filterwarnings('ignore')

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, 'data')
OUT_DIR  = os.path.join(ROOT, 'output')
PLOT_DIR = os.path.join(OUT_DIR, 'ml_plots_v2')
os.makedirs(PLOT_DIR, exist_ok=True)

DIAG_LABEL     = {0:'PD', 1:'PSP', 2:'MSA', 3:'CBS', 4:'DLB', 6:'ET', 7:'RBD', 10:'HC'}
AP_CODES       = {1, 2, 3, 4, 6, 7}


# ── Helpers ───────────────────────────────────────────────────────────────────
def load(name):
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    return df

def first_col(df, *keywords):
    kws = [k.lower() for k in keywords]
    for c in df.columns:
        if all(k in c.lower() for k in kws):
            return c
    return None

def yn_to_bin(v):
    """Convert a single Yes/No text value → 1/0/NaN."""
    s = str(v).lower()
    if 'yes' in s or 'oui' in s:
        return 1
    if 'no' in s or 'non' in s:
        return 0
    return np.nan

def save_fig(name):
    p = os.path.join(PLOT_DIR, name)
    plt.savefig(p, dpi=150, bbox_inches='tight')
    plt.close()


# ── Load forms ────────────────────────────────────────────────────────────────
print('Loading forms...')
enroll = load('Enrollement')
clin   = load('Clinical')
moca   = load('MoCA')
updrs  = load('MDS-UPDRS')
scopa  = load('SCOPA')
demo   = load('Demographic')
neuro  = load('Neuropsychological')


# ── Build label table from Clinical (Labelled) logic ─────────────────────────
print('Building labels...')

enrol_col    = first_col(enroll, 'nrolment', 'roup')
status_col   = first_col(enroll, 'Study Status')
complete_col = 'Complete?'   # literal column name in Enrollment CSV
diag_col     = first_col(clin, 'Determined diagnosis')

# Enrolled + Complete
enrolled_complete = enroll[
    enroll[status_col].str.contains('Enrolled', na=False) &
    (enroll[complete_col] == 'Complete')
].drop_duplicates('Project key')[['Project key', enrol_col]]

hc_keys = set(
    enrolled_complete[enrolled_complete[enrol_col].str.contains('Healthy', na=False)]['Project key']
)

# One label row per participant (deduplicate Clinical by project key)
clin_labels = clin[['Project key', diag_col]].drop_duplicates('Project key').copy()
clin_labels['det_code'] = pd.to_numeric(clin_labels[diag_col], errors='coerce')
# Assign HC = 10
clin_labels.loc[
    clin_labels['Project key'].isin(hc_keys) & clin_labels['det_code'].isna(), 'det_code'
] = 10
clin_labels['label'] = clin_labels['det_code'].map(DIAG_LABEL)

# Restrict to enrolled+complete
labels = enrolled_complete.merge(clin_labels[['Project key', 'det_code', 'label']],
                                  on='Project key', how='left')
print(f'  Enrolled+complete: {len(labels):,}')
print(f'  Label distribution:\n{labels["label"].value_counts(dropna=False).to_string()}')


# ── Feature engineering ───────────────────────────────────────────────────────
print('\nEngineering features...')

feats = labels[['Project key']].copy()

# ── 1. Clinical features ──────────────────────────────────────────────────────
clin_one = clin.drop_duplicates('Project key').copy()

# Numeric
dur_col    = first_col(clin_one, 'duration of disease at the time the clinical questionnaire')
onset_col  = first_col(clin_one, 'age at onset')
bmi_col    = first_col(clin_one, 'bmi')
height_col = first_col(clin_one, 'height of the participant (cm)')
weight_col = first_col(clin_one, 'weight of the participant (kg)')

for feat_name, col in [('disease_duration', dur_col), ('age_at_onset', onset_col),
                        ('bmi', bmi_col), ('height_cm', height_col), ('weight_kg', weight_col)]:
    if col:
        feats = feats.merge(
            clin_one[['Project key', col]].rename(columns={col: feat_name})
            .assign(**{feat_name: lambda d, c=col: pd.to_numeric(clin_one[c], errors='coerce')}),
            on='Project key', how='left'
        )

# Binary symptom flags
sym_map = {
    'has_dyskinesia':         first_col(clin_one, 'dyskinesia', 'currently'),
    'has_freezing':           first_col(clin_one, 'freezing of gait'),
    'has_falls':              first_col(clin_one, 'fallen in the last 3 months'),
    'has_dementia':           first_col(clin_one, 'does the patient have dementia'),
    'has_dbs':                first_col(clin_one, 'surgery for parkinson'),
    'has_motor_fluctuations': first_col(clin_one, 'motor fluctuations'),
    'has_remission':          first_col(clin_one, 'complete remission'),
    'bilateral_onset':        first_col(clin_one, 'both sides of your body'),
}
for feat_name, col in sym_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            clin_one.set_index('Project key')[col].apply(yn_to_bin)
        )

# Dominant hand
hand_col = first_col(clin_one, 'dominant hand')
if hand_col:
    feats['right_handed'] = feats['Project key'].map(
        clin_one.set_index('Project key')[hand_col]
        .apply(lambda v: 1 if 'right' in str(v).lower() else (0 if 'left' in str(v).lower() else np.nan))
    )

print(f'  Clinical features: {[c for c in feats.columns if c != "Project key"]}')

# ── 2. MoCA features ──────────────────────────────────────────────────────────
moca_one = moca.drop_duplicates('Project key').copy()
moca_score_map = {
    'moca_visuospatial': first_col(moca_one, 'visuospatial'),
    'moca_naming':       first_col(moca_one, 'naming', 'score'),
    'moca_attention':    first_col(moca_one, 'attention', 'score'),
    'moca_language':     first_col(moca_one, 'language', 'score'),
    'moca_abstraction':  first_col(moca_one, 'abstraction', 'score'),
    'moca_recall':       first_col(moca_one, 'recall', 'score'),
    'moca_orientation':  first_col(moca_one, 'orientation', 'score'),
    'moca_total':        first_col(moca_one, 'total score'),
}
for feat_name, col in moca_score_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            moca_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# ── 3. UPDRS Part totals ──────────────────────────────────────────────────────
updrs_one = updrs.drop_duplicates('Project key').copy()
updrs_map = {
    'updrs_part1': first_col(updrs_one, 'part i'),
    'updrs_part2': first_col(updrs_one, 'part ii'),
    'updrs_part3': first_col(updrs_one, 'part iii'),
    'updrs_part4': first_col(updrs_one, 'part iv'),
}
for feat_name, col in updrs_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            updrs_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# ── 4. SCOPA autonomic subscores ──────────────────────────────────────────────
scopa_one = scopa.drop_duplicates('Project key').copy()
scopa_map = {
    'scopa_gi':           first_col(scopa_one, 'gastrointestinal'),
    'scopa_urinary':      first_col(scopa_one, 'urinary'),
    'scopa_cardiovasc':   first_col(scopa_one, 'cardiovascular'),
    'scopa_thermoreg':    first_col(scopa_one, 'thermoregulatory'),
    'scopa_pupilmotor':   first_col(scopa_one, 'pupillomotor'),
}
for feat_name, col in scopa_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            scopa_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# ── 5. Demographics ───────────────────────────────────────────────────────────
demo_one = demo.drop_duplicates('Project key').copy()
age_col = first_col(demo_one, 'age at study visit', 'automatic') or first_col(demo_one, 'age at study visit')
sex_col = first_col(demo_one, 'gender')
edu_col = first_col(demo_one, 'years of education')

for feat_name, col in [('age', age_col), ('edu_years', edu_col)]:
    if col:
        feats[feat_name] = feats['Project key'].map(
            demo_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )
if sex_col:
    feats['sex_male'] = feats['Project key'].map(
        demo_one.set_index('Project key')[sex_col].apply(
            lambda v: 1 if 'male' in str(v).lower() and 'female' not in str(v).lower()
            else (0 if 'female' in str(v).lower() else np.nan)
        )
    )

# ── 6. Neuropsychological (key scores) ────────────────────────────────────────
neuro_one = neuro.drop_duplicates('Project key').copy()
neuro_map = {
    'neuro_hvlt_total':     first_col(neuro_one, 'trial total 1,2,3', 'raw'),
    'neuro_hvlt_delayed':   first_col(neuro_one, 'trial 4 delayed'),
    'neuro_bnt':            first_col(neuro_one, 'copy raw'),
    'neuro_digit_fwd':      first_col(neuro_one, 'digit span forward', 'total correct'),
    'neuro_digit_bwd':      first_col(neuro_one, 'digit span backward', 'total correct'),
    'neuro_trail_a':        first_col(neuro_one, 'trail a raw score'),
    'neuro_trail_b':        first_col(neuro_one, 'trail b raw score') or first_col(neuro_one, 'trail b'),
    'neuro_digit_sym':      first_col(neuro_one, 'digit'),
}
for feat_name, col in neuro_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            neuro_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# Final feature list — drop any column with zero coverage (useless for modelling)
FEAT_COLS = [
    c for c in feats.columns
    if c != 'Project key' and feats[c].notna().any()
]
print(f'\n  Total features: {len(FEAT_COLS)}')
print(f'  Features: {FEAT_COLS}')

# Merge labels
df = feats.merge(labels[['Project key', 'det_code', 'label']], on='Project key', how='inner')
print(f'\n  Master table: {len(df):,} rows × {len(FEAT_COLS)} features')


# ── Coverage report ───────────────────────────────────────────────────────────
print('\nFeature coverage (% non-null):')
for col in FEAT_COLS:
    pct = df[col].notna().mean() * 100
    print(f'  {col:30s}: {pct:5.1f}%')


# ── Model pipeline factories ──────────────────────────────────────────────────
def make_pipelines(cw=None):
    return {
        'Logistic Regression': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('scl', StandardScaler()),
            ('clf', LogisticRegression(max_iter=2000, class_weight=cw, random_state=42)),
        ]),
        'Random Forest': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('clf', RandomForestClassifier(n_estimators=500, class_weight=cw,
                                            random_state=42, n_jobs=-1)),
        ]),
        'XGBoost': Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('clf', XGBClassifier(n_estimators=500, learning_rate=0.05, max_depth=5,
                                   random_state=42, verbosity=0, eval_metric='logloss')),
        ]),
    }


# ── Classification runner ─────────────────────────────────────────────────────
all_results = []

def run_task(task_name, df_task, label_col, features, n_splits=5,
             cw=None, multiclass=False, xgb_scale_pos=None):
    print(f'\n{"="*60}')
    print(f'  {task_name}')
    print(f'{"="*60}')

    df_task = df_task[features + [label_col]].dropna(subset=[label_col]).copy()
    le = LabelEncoder()
    y  = le.fit_transform(df_task[label_col])
    X  = df_task[features].values
    classes = le.classes_

    counts = dict(zip(classes, np.bincount(y)))
    print(f'  N={len(y)}  |  {counts}')

    if len(np.unique(y)) < 2:
        print('  Skipping — only one class present.')
        return

    n_splits_eff = min(n_splits, int(np.bincount(y).min()))
    if n_splits_eff < 2:
        print('  Skipping — not enough samples for CV.')
        return
    cv = StratifiedKFold(n_splits=n_splits_eff, shuffle=True, random_state=42)

    pipelines = make_pipelines(cw=cw)
    # XGBoost scale_pos_weight for binary imbalanced
    if xgb_scale_pos and not multiclass:
        pipelines['XGBoost']['clf'].set_params(scale_pos_weight=xgb_scale_pos)

    best_auc, best_model_name, best_pipe = -1, None, None

    for model_name, pipe in pipelines.items():
        try:
            y_prob = cross_val_predict(pipe, X, y, cv=cv, method='predict_proba')
            y_pred = y_prob.argmax(axis=1)

            if multiclass:
                auc = roc_auc_score(y, y_prob, multi_class='ovr', average='macro')
            else:
                auc = roc_auc_score(y, y_prob[:, 1])

            bal_acc = balanced_accuracy_score(y, y_pred)
            rpt     = classification_report(y, y_pred, target_names=classes,
                                            output_dict=True, zero_division=0)
            row = {
                'Task':         task_name,
                'Model':        model_name,
                'N':            len(y),
                'Classes':      ' / '.join(str(c) for c in classes),
                'ROC-AUC':      round(auc, 3),
                'Balanced Acc': round(bal_acc, 3),
            }
            for cls in classes:
                row[f'Precision {cls}'] = round(rpt[cls]['precision'], 3)
                row[f'Recall {cls}']    = round(rpt[cls]['recall'],    3)
                row[f'F1 {cls}']        = round(rpt[cls]['f1-score'],  3)
            all_results.append(row)
            print(f'  {model_name:22s}  AUC={auc:.3f}  BalAcc={bal_acc:.3f}')

            if auc > best_auc:
                best_auc, best_model_name, best_pipe = auc, model_name, pipe

            # Confusion matrix
            if model_name == 'Random Forest':
                cm  = confusion_matrix(y, y_pred)
                fig, ax = plt.subplots(figsize=(max(4, len(classes)), max(3, len(classes))))
                sns.heatmap(cm, annot=True, fmt='d', xticklabels=classes,
                            yticklabels=classes, cmap='Blues', ax=ax)
                ax.set_title(f'{task_name}\n{model_name} — Confusion Matrix (CV)')
                ax.set_xlabel('Predicted'); ax.set_ylabel('Actual')
                plt.tight_layout()
                save_fig(re.sub(r'[^\w]', '_', f'{task_name}_{model_name}') + '_cm.png')

        except Exception as e:
            print(f'  {model_name}: ERROR — {e}')

    # Feature importance (Random Forest, full fit)
    try:
        rf = make_pipelines(cw=cw)['Random Forest']
        # Drop task-level all-NaN features before fitting
        valid_mask  = ~np.all(np.isnan(X.astype(float)), axis=0)
        X_valid     = X[:, valid_mask]
        feat_valid  = [f for f, m in zip(features, valid_mask) if m]
        X_imp = SimpleImputer(strategy='median').fit_transform(X_valid)
        rf['clf'].fit(X_imp, y)
        fi = pd.Series(rf['clf'].feature_importances_, index=feat_valid).sort_values()
        fig, ax = plt.subplots(figsize=(6, max(4, len(features) * 0.32)))
        fi.plot.barh(ax=ax, color='steelblue')
        ax.set_title(f'{task_name}\nRF Feature Importance')
        ax.axvline(fi.mean(), color='red', linestyle='--', linewidth=0.8, label='mean')
        ax.legend(fontsize=8)
        plt.tight_layout()
        save_fig(re.sub(r'[^\w]', '_', task_name) + '_feature_importance.png')
    except Exception as e:
        print(f'  Feature importance: {e}')

    # SHAP (best model, fit on full data)
    try:
        best_pipe_copy = make_pipelines(cw=cw)[best_model_name]
        valid_mask2 = ~np.all(np.isnan(X.astype(float)), axis=0)
        X_valid2    = X[:, valid_mask2]
        feat_valid2 = [f for f, m in zip(features, valid_mask2) if m]
        X_imp = SimpleImputer(strategy='median').fit_transform(X_valid2)
        best_pipe_copy['clf'].fit(X_imp, y)

        if 'XGBoost' in best_model_name or 'Random Forest' in best_model_name:
            explainer = shap.TreeExplainer(best_pipe_copy['clf'])
            shap_vals = explainer.shap_values(X_imp)
        else:
            explainer = shap.LinearExplainer(best_pipe_copy['clf'],
                                              StandardScaler().fit_transform(X_imp))
            shap_vals = explainer.shap_values(StandardScaler().fit_transform(X_imp))

        # For multiclass take mean abs across classes
        if isinstance(shap_vals, list):
            shap_arr = np.abs(np.array(shap_vals)).mean(axis=0)
        else:
            shap_arr = shap_vals if shap_vals.ndim == 2 else shap_vals

        plt.figure(figsize=(6, max(4, len(feat_valid2) * 0.32)))
        shap.summary_plot(shap_arr, X_imp, feature_names=feat_valid2,
                          plot_type='bar', show=False)
        plt.title(f'{task_name}\nSHAP — {best_model_name}')
        plt.tight_layout()
        save_fig(re.sub(r'[^\w]', '_', task_name) + '_shap.png')
        plt.close()
    except Exception as e:
        print(f'  SHAP: {e}')


# ═══════════════════════════════════════════════════════════════════════════════
# TASK A — PD vs HC
# ═══════════════════════════════════════════════════════════════════════════════
df_a = df[df['label'].isin(['PD', 'HC'])].copy()
run_task('Task A — PD vs HC', df_a, 'label', FEAT_COLS, cw='balanced')

# ═══════════════════════════════════════════════════════════════════════════════
# TASK B — PD vs AP
# ═══════════════════════════════════════════════════════════════════════════════
df_b = df[df['label'].isin(['PD']) | df['det_code'].isin(AP_CODES)].copy()
df_b['task_label'] = df_b['det_code'].apply(
    lambda c: 'PD' if c == 0 else ('AP' if c in AP_CODES else np.nan)
)
n_pd = (df_b['task_label'] == 'PD').sum()
n_ap = (df_b['task_label'] == 'AP').sum()
run_task('Task B — PD vs AP', df_b, 'task_label', FEAT_COLS,
         cw='balanced', xgb_scale_pos=round(n_pd / max(n_ap, 1)))

# ═══════════════════════════════════════════════════════════════════════════════
# TASK C — PD vs HC vs AP  (3-class)
# ═══════════════════════════════════════════════════════════════════════════════
df_c = df[df['label'].isin(['PD', 'HC']) | df['det_code'].isin(AP_CODES)].copy()
df_c['task_label'] = df_c['det_code'].apply(
    lambda c: 'PD' if c == 0 else ('HC' if c == 10 else ('AP' if c in AP_CODES else np.nan))
)
run_task('Task C — PD vs HC vs AP', df_c, 'task_label', FEAT_COLS,
         cw='balanced', multiclass=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TASK D — AP subtype (PSP / MSA / DLB+other) — exploratory
# ═══════════════════════════════════════════════════════════════════════════════
df_d = df[df['det_code'].isin(AP_CODES)].copy()
df_d['task_label'] = df_d['det_code'].apply(
    lambda c: 'PSP' if c == 1 else ('MSA' if c == 2 else 'DLB+other')
)
run_task('Task D — AP subtype (exploratory)', df_d, 'task_label', FEAT_COLS,
         n_splits=3, cw='balanced', multiclass=True)


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
print('\n\nWriting results...')
results_df = pd.DataFrame(all_results)

# Coverage table
coverage = pd.DataFrame({
    'Feature':   FEAT_COLS,
    'Non-null N': [int(df[c].notna().sum()) for c in FEAT_COLS],
    'Coverage %': [round(df[c].notna().mean() * 100, 1) for c in FEAT_COLS],
    'Group':      [
        'Clinical' if c in ['disease_duration','age_at_onset','bmi','height_cm','weight_kg',
                             'has_dyskinesia','has_freezing','has_falls','has_dementia',
                             'has_dbs','has_motor_fluctuations','has_remission',
                             'bilateral_onset','right_handed']
        else 'MoCA' if c.startswith('moca')
        else 'UPDRS' if c.startswith('updrs')
        else 'SCOPA' if c.startswith('scopa')
        else 'Demographic' if c in ['age','edu_years','sex_male']
        else 'Neuropsychological'
        for c in FEAT_COLS
    ],
})

notes = pd.DataFrame([
    ['Task A', 'PD vs HC',           'Binary',     f'{(df_a["label"]=="PD").sum()} PD / {(df_a["label"]=="HC").sum()} HC',  'class_weight=balanced'],
    ['Task B', 'PD vs AP',           'Binary',     f'{n_pd} PD / {n_ap} AP',                                                 'class_weight=balanced + XGB scale_pos_weight'],
    ['Task C', 'PD vs HC vs AP',     '3-class',    'PD / HC / AP combined',                                                  'class_weight=balanced'],
    ['Task D', 'AP subtype',         'Multiclass', 'PSP / MSA / DLB+other  — EXPLORATORY, N~100',                           '3-fold CV only, do not overinterpret'],
], columns=['Task', 'Description', 'Type', 'Classes/N', 'Notes'])

out_path = os.path.join(OUT_DIR, 'ml_results_v2.xlsx')
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils  import get_column_letter
import openpyxl

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    results_df.to_excel(writer, sheet_name='Results',  index=False)
    coverage.to_excel(writer,   sheet_name='Feature Coverage', index=False)
    notes.to_excel(writer,      sheet_name='Task Notes', index=False)

wb = openpyxl.load_workbook(out_path)
HDR = PatternFill('solid', fgColor='1F4E79')
HF  = Font(color='FFFFFF', bold=True)

TASK_FILLS = {
    'Task A': PatternFill('solid', fgColor='DFF0D8'),
    'Task B': PatternFill('solid', fgColor='FFF2CC'),
    'Task C': PatternFill('solid', fgColor='D9E1F2'),
    'Task D': PatternFill('solid', fgColor='FCE4D6'),
}
GROUP_FILLS = {
    'Clinical':          PatternFill('solid', fgColor='DFF0D8'),
    'MoCA':              PatternFill('solid', fgColor='D9E1F2'),
    'UPDRS':             PatternFill('solid', fgColor='FFF2CC'),
    'SCOPA':             PatternFill('solid', fgColor='FCE4D6'),
    'Demographic':       PatternFill('solid', fgColor='EAD1DC'),
    'Neuropsychological':PatternFill('solid', fgColor='E2EFDA'),
}

for sn in wb.sheetnames:
    ws = wb[sn]
    for cell in ws[1]:
        cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 45)

ws_r = wb['Results']
task_col = next((i for i, c in enumerate(ws_r[1], 1) if c.value == 'Task'), None)
if task_col:
    for row in ws_r.iter_rows(min_row=2):
        task_val = str(row[task_col - 1].value or '')
        fill = next((f for k, f in TASK_FILLS.items() if k in task_val), None)
        if fill:
            for cell in row: cell.fill = fill

ws_c = wb['Feature Coverage']
grp_col = next((i for i, c in enumerate(ws_c[1], 1) if c.value == 'Group'), None)
if grp_col:
    for row in ws_c.iter_rows(min_row=2):
        grp = str(row[grp_col - 1].value or '')
        fill = GROUP_FILLS.get(grp)
        if fill:
            for cell in row: cell.fill = fill

wb.save(out_path)
print(f'  Results → {out_path}')
print(f'  Plots   → {PLOT_DIR}/')

print('\n── Summary ──────────────────────────────────────────────────────')
print(results_df[['Task', 'Model', 'N', 'ROC-AUC', 'Balanced Acc']].to_string(index=False))
