"""
build_updrs.py
Creates two derived UPDRS forms from the three source forms:

  UPDRS_general  — MDS-UPDRS Parts 1, 2, 3-admin, 4 + HY  (MDS-UPDRS + MDS-UPDRS-1)
  UPDRS_part3    — MDS-UPDRS Part 3 motor exam             (all three source forms)

For participants with multiple entries in either derived form, a delta sheet is
produced with time-between-assessments and score differences per numeric variable.

Outputs:  updrs_analysis.xlsx
Run:      python3 build_updrs.py
"""

import os
import re
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, 'data')
OUT_PATH = os.path.join(ROOT, 'output', 'updrs_analysis.xlsx')

ADMIN_COLS = {'Project key', 'Event Name', 'Complete?', 'Source Form'}

# ── Load raw source forms ──────────────────────────────────────────────────────
def load_csv(name: str) -> pd.DataFrame:
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    return df

mds  = load_csv('MDS-UPDRS')
mds1 = load_csv('MDS-UPDRS-1')
u12  = load_csv('UPDRS 1.2 part 3')

MC = list(mds.columns)   # canonical MDS-UPDRS column names
UC = list(u12.columns)   # UPDRS 1.2 column names

print(f'Loaded: MDS-UPDRS ({len(mds)} rows), MDS-UPDRS-1 ({len(mds1)} rows), '
      f'UPDRS 1.2 part 3 ({len(u12)} rows)')

# ── Column index groups (positions in MC / MDS-UPDRS) ─────────────────────────
# UPDRS_general: admin + Part 1 + Part 2 + Part 3-admin + dyskinesias/HY + Part 4
GEN_IDX = (
    list(range(0, 46))    +   # admin (0-6), Part1 (7-26), Part2 (27-39), Part3-admin (40-45)
    list(range(112, 129)) +   # dyskinesias (112-113), HY (114), Part4 items (115-128)
    [130, 131]                # Part IV total, Complete?
)

# UPDRS_part3: admin + Part 3 items + Part III total
P3_IDX = (
    list(range(0, 7))    +   # admin
    list(range(46, 112)) +   # Part 3 label+value cols (46-111)
    [129, 131]               # Part III total, Complete?
)

# ── UPDRS_1.2 → MDS column mapping for UPDRS_part3 ────────────────────────────
# Each entry: UC[i] → (MDS label col idx, MDS value col idx)
# Label col contains text; value col (label+1) holds the numeric score.
U12_TO_MDS_PAIR = {
    UC[4]:  (46,  47),   # 1. Speech           → 3.1 SPEECH / Updrs_3_1 value
    UC[5]:  (48,  49),   # 2. Facial Express   → 3.2 FACIAL EXPRESSION / Updrs_3_2 value
    UC[6]:  (108, 109),  # 3a. face/lips/chin  → 3.17 REST Lip/Jaw / Updrs_3_17_lipjaw value
    UC[7]:  (100, 101),  # 3b. R hand          → 3.17 REST RUE / Updrs_3_17_rue value
    UC[8]:  (102, 103),  # 3c. L hand          → 3.17 REST LUE / Updrs_3_17_lue value
    UC[9]:  (104, 105),  # 3d. R foot          → 3.17 REST RLE / Updrs_3_17_rle value
    UC[10]: (106, 107),  # 3e. L foot          → 3.17 REST LLE / Updrs_3_17_lle value
    UC[11]: (96,  97),   # 4a. kinetic R       → 3.16 KINETIC R / Updrs_3_16_r value
    UC[12]: (98,  99),   # 4b. kinetic L       → 3.16 KINETIC L / Updrs_3_16_l value
    UC[13]: (50,  51),   # 5a. rigidity neck   → 3.3 RIGIDITY NECK / Updrs_3_3_neck value
    UC[14]: (52,  53),   # 5b. rigidity RUE    → 3.3 RIGIDITY RUE / Updrs_3_3_rue value
    UC[15]: (54,  55),   # 5c. rigidity LUE    → 3.3 RIGIDITY LUE / Updrs_3_3_lue value
    UC[16]: (56,  57),   # 5d. rigidity RLE    → 3.3 RIGIDITY RLE / Updrs_3_3_rle value
    UC[17]: (58,  59),   # 5e. rigidity LLE    → 3.3 RIGIDITY LLE / Updrs_3_3_lle value
    UC[18]: (60,  61),   # 6a. finger taps R   → 3.4 FINGER TAPPING R / Updrs_3_4_r value
    UC[19]: (62,  63),   # 6b. finger taps L   → 3.4 FINGER TAPPING L / Updrs_3_4_l value
    UC[20]: (64,  65),   # 7a. hand mov R      → 3.5 HAND MOVEMENTS R / Updrs_3_5_r value
    UC[21]: (66,  67),   # 7b. hand mov L      → 3.5 HAND MOVEMENTS L / Updrs_3_5_l value
    UC[22]: (68,  69),   # 8a. alt hand R      → 3.6 PRON-SUP R / Updrs_3_6_r value
    UC[23]: (70,  71),   # 8b. alt hand L      → 3.6 PRON-SUP L / Updrs_3_6_l value
    UC[24]: (76,  77),   # 9a. leg agility R   → 3.8 LEG AGILITY R / Updrs_3_8_r value
    UC[25]: (78,  79),   # 9b. leg agility L   → 3.8 LEG AGILITY L / Updrs_3_8_l value
    UC[26]: (80,  81),   # 10. arise chair     → 3.9 ARISING FROM CHAIR / Updrs_3_9 value
    UC[27]: (88,  89),   # 11. posture         → 3.13 POSTURE / Updrs_3_13 value
    UC[28]: (82,  83),   # 12. gait            → 3.10 GAIT / Updrs_3_10 value
    UC[29]: (86,  87),   # 13. postural stab   → 3.12 POSTURAL STABILITY / Updrs_3_12 value
    UC[30]: (90,  91),   # 14. body brady      → 3.14 GLOBAL SPONT / Updrs_3_14
}

# Admin mapping: UC[i] → MC[j]
U12_ADMIN_MAP = {
    UC[0]:  MC[0],   # Project key
    UC[1]:  MC[1],   # Event Name
    UC[2]:  MC[2],   # Assessment completed
    UC[3]:  MC[4],   # How administered
    UC[38]: MC[131], # Complete?
}

# UPDRS_1.2-only total columns (appended to UPDRS_part3; NaN for MDS rows)
U12_TOTAL_COLS = [UC[31], UC[32], UC[33], UC[34], UC[35], UC[36], UC[37]]
# UPDRS Total, Tremor Total, Rigidity Total, Ratio T/R, Right, Left, Laterality

PART3_COLS = ['Source Form'] + [MC[i] for i in P3_IDX] + U12_TOTAL_COLS
GEN_COLS   = ['Source Form'] + [MC[i] for i in GEN_IDX]


def extract_score(val) -> float:
    """Parse leading integer from UPDRS 1.2 text like '1 - Slight...' or '0 - Normal'."""
    if pd.isna(val):
        return np.nan
    m = re.match(r'^\s*(\d+)', str(val))
    return float(m.group(1)) if m else np.nan


# ── Build UPDRS_general ────────────────────────────────────────────────────────
print('Building UPDRS_general...')

def make_general(df: pd.DataFrame, source: str) -> pd.DataFrame:
    sub = df.iloc[:, GEN_IDX].copy()
    sub.insert(0, 'Source Form', source)
    sub.columns = GEN_COLS
    return sub

updrs_general = pd.concat([
    make_general(mds,  'MDS-UPDRS'),
    make_general(mds1, 'MDS-UPDRS-1'),
], ignore_index=True)

data_g = [c for c in updrs_general.columns if c not in ADMIN_COLS]
updrs_general = updrs_general[updrs_general[data_g].notna().any(axis=1)].reset_index(drop=True)
print(f'  UPDRS_general: {len(updrs_general)} rows  '
      f'({(updrs_general["Source Form"]=="MDS-UPDRS").sum()} MDS-UPDRS, '
      f'{(updrs_general["Source Form"]=="MDS-UPDRS-1").sum()} MDS-UPDRS-1)')


# ── Build UPDRS_part3 ──────────────────────────────────────────────────────────
print('Building UPDRS_part3...')

def make_part3_mds(df: pd.DataFrame, source: str) -> pd.DataFrame:
    sub = df.iloc[:, P3_IDX].copy()
    sub.columns = [MC[i] for i in P3_IDX]
    sub.insert(0, 'Source Form', source)
    for c in U12_TOTAL_COLS:
        sub[c] = np.nan
    return sub[PART3_COLS]

def make_part3_u12(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(np.nan, index=df.index, columns=PART3_COLS)
    out['Source Form'] = 'UPDRS 1.2 part 3'

    # Admin
    for u12_col, mds_col in U12_ADMIN_MAP.items():
        if u12_col in df.columns:
            out[mds_col] = df[u12_col].values

    # Score items: text → label column, parsed numeric → value column
    for u12_col, (label_idx, value_idx) in U12_TO_MDS_PAIR.items():
        if u12_col not in df.columns:
            continue
        label_col = MC[label_idx]
        value_col = MC[value_idx]
        if label_col in out.columns:
            out[label_col] = df[u12_col].values                     # text label
        if value_col in out.columns:
            out[value_col] = df[u12_col].apply(extract_score).values  # numeric

    # UPDRS_1.2 total columns
    for u12_col in U12_TOTAL_COLS:
        if u12_col in df.columns:
            out[u12_col] = pd.to_numeric(df[u12_col], errors='coerce').values

    return out

updrs_part3 = pd.concat([
    make_part3_mds(mds,  'MDS-UPDRS'),
    make_part3_mds(mds1, 'MDS-UPDRS-1'),
    make_part3_u12(u12),
], ignore_index=True)

data_p3 = [c for c in updrs_part3.columns if c not in ADMIN_COLS]
updrs_part3 = updrs_part3[updrs_part3[data_p3].notna().any(axis=1)].reset_index(drop=True)
src_counts = updrs_part3['Source Form'].value_counts()
print(f'  UPDRS_part3: {len(updrs_part3)} rows  ' +
      '  '.join(f'({v} {k})' for k, v in src_counts.items()))


# ── Multi-entry delta analysis ─────────────────────────────────────────────────
print('Computing multi-entry deltas...')

DATE_COL = MC[2]   # 'Assessment completed:     Évaluation remplie:  '

def compute_deltas(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """
    For each participant with >1 entry, compute consecutive-pair differences.
    Returns one row per consecutive (earlier, later) pair.
    """
    df = df.copy()
    df[DATE_COL] = pd.to_datetime(df[DATE_COL], errors='coerce')

    # Identify numeric columns (value cols, totals — excludes text label cols)
    num_cols = [
        c for c in df.columns
        if c not in ADMIN_COLS and c != DATE_COL
        and pd.api.types.is_numeric_dtype(df[c])
    ]

    delta_rows = []
    multi_pids = df.groupby('Project key').filter(lambda g: len(g) > 1)

    for pid, grp in multi_pids.groupby('Project key'):
        grp = grp.sort_values(DATE_COL, na_position='last').reset_index(drop=True)
        for i in range(len(grp) - 1):
            r1, r2 = grp.iloc[i], grp.iloc[i + 1]
            d1, d2 = r1[DATE_COL], r2[DATE_COL]
            days = int((d2 - d1).days) if pd.notna(d1) and pd.notna(d2) else np.nan

            row: dict = {
                'Project key':   pid,
                'Source Form 1': r1['Source Form'],
                'Source Form 2': r2['Source Form'],
                'Date 1':        d1.date() if pd.notna(d1) else np.nan,
                'Date 2':        d2.date() if pd.notna(d2) else np.nan,
                'Days Between':  days,
            }
            for c in num_cols:
                v1 = pd.to_numeric(r1[c], errors='coerce')
                v2 = pd.to_numeric(r2[c], errors='coerce')
                if pd.notna(v1) or pd.notna(v2):
                    row[f'Δ {c}'] = (v2 - v1) if (pd.notna(v1) and pd.notna(v2)) else np.nan

            delta_rows.append(row)

    result = pd.DataFrame(delta_rows)
    n_pids = multi_pids['Project key'].nunique()
    print(f'  {label}: {n_pids} participants with multiple entries → '
          f'{len(result)} consecutive pairs')
    return result

delta_general = compute_deltas(updrs_general, 'UPDRS_general')
delta_part3   = compute_deltas(updrs_part3,   'UPDRS_part3')


# ── Formatting helpers ─────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='212529')
HDR_FONT  = Font(color='FFFFFF', bold=True, size=11)
EVEN_FILL = PatternFill('solid', fgColor='F8F9FA')
ODD_FILL  = PatternFill('solid', fgColor='FFFFFF')
POS_FILL  = PatternFill('solid', fgColor='C6EFCE')   # green: improvement (Δ < 0)
NEG_FILL  = PatternFill('solid', fgColor='FFC7CE')   # red:   worsening  (Δ > 0)

SOURCE_FILLS = {
    'MDS-UPDRS':        PatternFill('solid', fgColor='D0E8FF'),
    'MDS-UPDRS-1':      PatternFill('solid', fgColor='BDD7EE'),
    'UPDRS 1.2 part 3': PatternFill('solid', fgColor='FFE8CC'),
}

def autowidth(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 55)

def style_form_sheet(ws, source_col_idx: int = 1):
    """Header + zebra rows + Source Form colour coding."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        src = row[source_col_idx - 1].value or ''
        rf  = SOURCE_FILLS.get(src, EVEN_FILL if i % 2 == 0 else ODD_FILL)
        for cell in row:
            cell.fill = rf
    autowidth(ws)

def style_delta_sheet(ws, delta_start_col: int):
    """Header + zebra rows + green/red on delta columns."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
        for cell in row[delta_start_col - 1:]:
            try:
                v = float(cell.value)
                if v < 0:
                    cell.fill = POS_FILL  # score decreased = improvement
                elif v > 0:
                    cell.fill = NEG_FILL  # score increased = worsening
            except (ValueError, TypeError):
                pass
    autowidth(ws)


# ── Write output ──────────────────────────────────────────────────────────────
print(f'Writing {OUT_PATH}...')

with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    updrs_general.to_excel(writer, sheet_name='UPDRS_general', index=False)
    updrs_part3.to_excel(writer,   sheet_name='UPDRS_part3',   index=False)
    if len(delta_general) > 0:
        delta_general.to_excel(writer, sheet_name='UPDRS_general_delta', index=False)
    if len(delta_part3) > 0:
        delta_part3.to_excel(writer,   sheet_name='UPDRS_part3_delta',   index=False)

wb = openpyxl.load_workbook(OUT_PATH)

# Source Form is column 1 in both form sheets
style_form_sheet(wb['UPDRS_general'], source_col_idx=1)
style_form_sheet(wb['UPDRS_part3'],   source_col_idx=1)

# Delta sheets: first 6 cols are metadata; delta values start at col 7
if 'UPDRS_general_delta' in wb.sheetnames:
    style_delta_sheet(wb['UPDRS_general_delta'], delta_start_col=7)
if 'UPDRS_part3_delta' in wb.sheetnames:
    style_delta_sheet(wb['UPDRS_part3_delta'],   delta_start_col=7)

wb.save(OUT_PATH)

print(f'\nDone. Sheets in {OUT_PATH}:')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row - 1:,} rows × {ws.max_column} cols')
