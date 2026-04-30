"""
build_ml_v3.py
Diagnostic classification pipeline — NO imputation.

Missing values are handled by each model natively:
  XGBoost                    — learns optimal split direction for NaN
  HistGradientBoosting       — sklearn's NaN-native gradient boosting
  Logistic Regression        — complete cases only, features restricted
                               to those with ≥ MIN_COVERAGE coverage

This lets you see what the data actually says, without median fill-in
artificially flattening variance in sparse features.

Label map: 0=PD  1=PSP  2=MSA  3=CBS  4=DLB  6=ET  7=RBD  10=HC

Tasks
-----
  A — PD vs HC
  B — PD vs AP  (ET excluded)
  C — PD vs HC vs AP  (3-class)
  D — AP subtype: PSP / MSA / DLB+other  (exploratory)
  E — HC vs Non-HC

Output
------
  output/ml_results_v3.xlsx
  output/ml_plots_v3/

Run: python3 scripts/build_ml_v3.py
"""

import os, re, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import shap

from sklearn.pipeline        import Pipeline
from sklearn.preprocessing   import StandardScaler, LabelEncoder
from sklearn.linear_model    import LogisticRegression
from sklearn.ensemble        import HistGradientBoostingClassifier
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics         import (
    roc_auc_score, balanced_accuracy_score,
    classification_report, confusion_matrix,
)
from xgboost import XGBClassifier

warnings.filterwarnings('ignore')

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, 'data')
OUT_DIR  = os.path.join(ROOT, 'output')
PLOT_DIR = os.path.join(OUT_DIR, 'ml_plots_v3')
os.makedirs(PLOT_DIR, exist_ok=True)

# Minimum feature coverage to include in the LR complete-case model
MIN_COVERAGE = 0.80

DIAG_LABEL     = {0:'PD', 1:'PSP', 2:'MSA', 3:'CBS', 4:'DLB', 6:'ET', 7:'RBD', 10:'HC'}
AP_CODES       = {1, 2, 3, 4, 6, 7}
AP_CODES_NO_ET = {1, 2, 3, 4, 7}


# ── Helpers ───────────────────────────────────────────────────────────────────
def load(name):
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    return df

def first_col(df, *keywords):
    kws = [k.lower() for k in keywords]
    for c in df.columns:
        if all(k in c.lower() for k in kws):
            return c
    return None

def yn_to_bin(v):
    s = str(v).lower()
    if 'yes' in s or 'oui' in s:
        return 1
    if 'no' in s or 'non' in s:
        return 0
    return np.nan

def save_fig(name):
    p = os.path.join(PLOT_DIR, name)
    plt.savefig(p, dpi=150, bbox_inches='tight')
    plt.close()


# ── Load forms ────────────────────────────────────────────────────────────────
print('Loading forms...')
enroll = load('Enrollement')
clin   = load('Clinical')
moca   = load('MoCA')
updrs  = load('MDS-UPDRS')
updrs3 = load('UPDRS 1.2 part 3')
med    = load('Medication')
scopa  = load('SCOPA')
demo   = load('Demographic')
neuro  = load('Neuropsychological')


# ── Build label table ─────────────────────────────────────────────────────────
print('Building labels...')

enrol_col    = first_col(enroll, 'nrolment', 'roup')
status_col   = first_col(enroll, 'Study Status')
complete_col = 'Complete?'
diag_col     = first_col(clin, 'Determined diagnosis')

enrolled_complete = enroll[
    enroll[status_col].str.contains('Enrolled', na=False) &
    (enroll[complete_col] == 'Complete')
].drop_duplicates('Project key')[['Project key', enrol_col]]

hc_keys = set(
    enrolled_complete[
        enrolled_complete[enrol_col].str.contains('Healthy', na=False)
    ]['Project key']
)

clin_labels = clin[['Project key', diag_col]].drop_duplicates('Project key').copy()
clin_labels['det_code'] = pd.to_numeric(clin_labels[diag_col], errors='coerce')
clin_labels.loc[
    clin_labels['Project key'].isin(hc_keys) & clin_labels['det_code'].isna(), 'det_code'
] = 10
clin_labels['label'] = clin_labels['det_code'].map(DIAG_LABEL)

labels = enrolled_complete.merge(
    clin_labels[['Project key', 'det_code', 'label']], on='Project key', how='left'
)
print(f'  Enrolled+complete: {len(labels):,}')
print(f'  Label distribution:\n{labels["label"].value_counts(dropna=False).to_string()}')


# ── Feature engineering (identical to v2) ─────────────────────────────────────
print('\nEngineering features...')
feats = labels[['Project key']].copy()

# 1. Clinical
clin_one = clin.drop_duplicates('Project key').copy()
dur_col    = first_col(clin_one, 'duration of disease at the time the clinical questionnaire')
onset_col  = first_col(clin_one, 'age at onset')
bmi_col    = first_col(clin_one, 'bmi')
height_col = first_col(clin_one, 'height of the participant (cm)')
weight_col = first_col(clin_one, 'weight of the participant (kg)')

for feat_name, col in [('disease_duration', dur_col), ('age_at_onset', onset_col),
                        ('bmi', bmi_col), ('height_cm', height_col), ('weight_kg', weight_col)]:
    if col:
        feats = feats.merge(
            clin_one[['Project key', col]].rename(columns={col: feat_name})
            .assign(**{feat_name: lambda d, c=col: pd.to_numeric(clin_one[c], errors='coerce')}),
            on='Project key', how='left'
        )

sym_map = {
    'has_dyskinesia':         first_col(clin_one, 'dyskinesia', 'currently'),
    'has_freezing':           first_col(clin_one, 'freezing of gait'),
    'has_falls':              first_col(clin_one, 'fallen in the last 3 months'),
    'has_dementia':           first_col(clin_one, 'does the patient have dementia'),
    'has_dbs':                first_col(clin_one, 'surgery for parkinson'),
    'has_motor_fluctuations': first_col(clin_one, 'motor fluctuations'),
    'has_remission':          first_col(clin_one, 'complete remission'),
    'bilateral_onset':        first_col(clin_one, 'both sides of your body'),
}
for feat_name, col in sym_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            clin_one.set_index('Project key')[col].apply(yn_to_bin)
        )

hand_col = first_col(clin_one, 'dominant hand')
if hand_col:
    feats['right_handed'] = feats['Project key'].map(
        clin_one.set_index('Project key')[hand_col].apply(
            lambda v: 1 if 'right' in str(v).lower() else (0 if 'left' in str(v).lower() else np.nan)
        )
    )

# 2. MoCA
moca_one = moca.drop_duplicates('Project key').copy()
moca_score_map = {
    'moca_visuospatial': first_col(moca_one, 'visuospatial'),
    'moca_naming':       first_col(moca_one, 'naming', 'score'),
    'moca_attention':    first_col(moca_one, 'attention', 'score'),
    'moca_language':     first_col(moca_one, 'language', 'score'),
    'moca_abstraction':  first_col(moca_one, 'abstraction', 'score'),
    'moca_recall':       first_col(moca_one, 'recall', 'score'),
    'moca_orientation':  first_col(moca_one, 'orientation', 'score'),
    'moca_total':        first_col(moca_one, 'total score'),
}
for feat_name, col in moca_score_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            moca_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# 3. UPDRS Part totals
updrs_one = updrs.drop_duplicates('Project key').copy()
updrs_map = {
    'updrs_part1': first_col(updrs_one, 'part i'),
    'updrs_part2': first_col(updrs_one, 'part ii'),
    'updrs_part3': first_col(updrs_one, 'part iii'),
    'updrs_part4': first_col(updrs_one, 'part iv'),
}
for feat_name, col in updrs_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            updrs_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# 4. SCOPA
scopa_one = scopa.drop_duplicates('Project key').copy()
scopa_map = {
    'scopa_gi':         first_col(scopa_one, 'gastrointestinal'),
    'scopa_urinary':    first_col(scopa_one, 'urinary'),
    'scopa_cardiovasc': first_col(scopa_one, 'cardiovascular'),
    'scopa_thermoreg':  first_col(scopa_one, 'thermoregulatory'),
    'scopa_pupilmotor': first_col(scopa_one, 'pupillomotor'),
}
for feat_name, col in scopa_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            scopa_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# 5. Demographics
demo_one = demo.drop_duplicates('Project key').copy()
age_col = first_col(demo_one, 'age at study visit', 'automatic') or first_col(demo_one, 'age at study visit')
sex_col = first_col(demo_one, 'gender')
edu_col = first_col(demo_one, 'years of education')

for feat_name, col in [('age', age_col), ('edu_years', edu_col)]:
    if col:
        feats[feat_name] = feats['Project key'].map(
            demo_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )
if sex_col:
    feats['sex_male'] = feats['Project key'].map(
        demo_one.set_index('Project key')[sex_col].apply(
            lambda v: 1 if 'male' in str(v).lower() and 'female' not in str(v).lower()
            else (0 if 'female' in str(v).lower() else np.nan)
        )
    )

# 6. Neuropsychological
neuro_one = neuro.drop_duplicates('Project key').copy()
neuro_map = {
    'neuro_hvlt_total':   first_col(neuro_one, 'trial total 1,2,3', 'raw'),
    'neuro_hvlt_delayed': first_col(neuro_one, 'trial 4 delayed'),
    'neuro_bnt':          first_col(neuro_one, 'copy raw'),
    'neuro_digit_fwd':    first_col(neuro_one, 'digit span forward', 'total correct'),
    'neuro_digit_bwd':    first_col(neuro_one, 'digit span backward', 'total correct'),
    'neuro_trail_a':      first_col(neuro_one, 'trail a raw score'),
    'neuro_trail_b':      first_col(neuro_one, 'trail b raw score') or first_col(neuro_one, 'trail b'),
    'neuro_digit_sym':    first_col(neuro_one, 'digit'),
}
for feat_name, col in neuro_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            neuro_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

# 7. UPDRS 1.2 Part 3 subscores
updrs3_one = updrs3.drop_duplicates('Project key').copy()
updrs3_map = {
    'u3_tremor_total':       first_col(updrs3_one, 'tremor total'),
    'u3_rigidity_total':     first_col(updrs3_one, 'rigidity total'),
    'u3_tremor_rigid_ratio': first_col(updrs3_one, 'ratio tremor'),
    'u3_right_score':        first_col(updrs3_one, 'right part score'),
    'u3_left_score':         first_col(updrs3_one, 'left part score'),
    'u3_laterality_index':   first_col(updrs3_one, 'laterality index'),
}
for feat_name, col in updrs3_map.items():
    if col:
        feats[feat_name] = feats['Project key'].map(
            updrs3_one.set_index('Project key')[col].apply(pd.to_numeric, errors='coerce')
        )

_r_col = updrs3_map.get('u3_right_score')
_l_col = updrs3_map.get('u3_left_score')
if _r_col and _l_col:
    _r = updrs3_one.set_index('Project key')[_r_col].apply(pd.to_numeric, errors='coerce')
    _l = updrs3_one.set_index('Project key')[_l_col].apply(pd.to_numeric, errors='coerce')
    _r = feats['Project key'].map(_r)
    _l = feats['Project key'].map(_l)
    _s = _r + _l
    feats['u3_asymmetry'] = np.where(_s > 0, (_r - _l).abs() / _s, np.nan)

# 8. MDS-UPDRS PIGD / Tremor-Dominant subscores
updrs_one = updrs.drop_duplicates('Project key').copy()
_pigd_cols   = [c for c in updrs_one.columns if
                any(f'3_{n} ' in c or f'3_{n}\n' in c or c.strip().endswith(f'3_{n}')
                    or f'Updrs_3_{n}' in c for n in ('10','11','12','13','14'))]
_tremor_cols = [c for c in updrs_one.columns if
                any(f'Updrs_3_{n}' in c for n in ('15','16','17','18'))]
if _pigd_cols:
    feats['updrs_pigd'] = feats['Project key'].map(
        updrs_one.set_index('Project key')[_pigd_cols]
        .apply(pd.to_numeric, errors='coerce').mean(axis=1)
    )
if _tremor_cols:
    feats['updrs_tremor_items'] = feats['Project key'].map(
        updrs_one.set_index('Project key')[_tremor_cols]
        .apply(pd.to_numeric, errors='coerce').mean(axis=1)
    )
if _pigd_cols and _tremor_cols:
    _t = feats['updrs_tremor_items'].fillna(0)
    _p = feats['updrs_pigd'].fillna(0)
    feats['updrs_td_score'] = np.where((_t + _p) > 0, _t / (_t + _p), np.nan)

# 9. Medication — levodopa
med_one        = med.drop_duplicates('Project key').copy()
levo_resp_col  = first_col(med_one, 'significant reduction')
levo_cur_col   = first_col(med_one, 'Select all current')
levo_dose_cols = [c for c in med_one.columns
                  if 'levodopa dosage' in c.lower() or 'lévodopa' in c.lower()]

if levo_resp_col:
    feats['levo_response'] = feats['Project key'].map(
        med_one.set_index('Project key')[levo_resp_col].apply(
            lambda v: 1.0 if 'yes' in str(v).lower() or 'oui' in str(v).lower()
            else (0.0 if 'no/' in str(v).lower() or 'non/' in str(v).lower()
                  else (0.5 if 'uncertain' in str(v).lower() else np.nan))
        )
    )
if levo_cur_col:
    feats['on_levodopa'] = feats['Project key'].map(
        med_one.set_index('Project key')[levo_cur_col].apply(
            lambda v: 1 if 'levodopa' in str(v).lower() else (0 if pd.notna(v) else np.nan)
        )
    )
if levo_dose_cols:
    _dose = med_one.set_index('Project key')[levo_dose_cols].apply(
        pd.to_numeric, errors='coerce'
    ).sum(axis=1, min_count=1)
    feats['total_levo_dose'] = feats['Project key'].map(_dose)

# All features with at least some data
FEAT_COLS = [c for c in feats.columns if c != 'Project key' and feats[c].notna().any()]
print(f'  Total features: {len(FEAT_COLS)}')

df = feats.merge(labels[['Project key', 'det_code', 'label']], on='Project key', how='inner')
print(f'  Master table: {len(df):,} rows × {len(FEAT_COLS)} features')

# Coverage per feature (computed once on the full master table)
coverage_pct = {c: df[c].notna().mean() for c in FEAT_COLS}
print('\nFeature coverage (% non-null):')
for c, pct in coverage_pct.items():
    print(f'  {c:30s}: {pct*100:5.1f}%')


# ── Model factories — NO imputation ──────────────────────────────────────────
#
# XGBoost:           pass NaN directly; XGBoost learns the best split direction
# HistGradientBoost: sklearn's NaN-native tree ensemble
# Logistic Reg:      complete cases only on features with ≥ MIN_COVERAGE coverage
#
def make_xgb(cw=None, scale_pos=None):
    params = dict(n_estimators=500, learning_rate=0.05, max_depth=5,
                  random_state=42, verbosity=0, eval_metric='logloss',
                  tree_method='hist')  # hist supports NaN natively
    if scale_pos:
        params['scale_pos_weight'] = scale_pos
    return XGBClassifier(**params)

def make_hgb(cw=None):
    return HistGradientBoostingClassifier(
        max_iter=500, random_state=42,
        class_weight=cw,   # 'balanced' or None
    )

def make_lr(cw=None):
    return Pipeline([
        ('scl', StandardScaler()),
        ('clf', LogisticRegression(max_iter=2000, class_weight=cw, random_state=42)),
    ])


# ── Classification runner ─────────────────────────────────────────────────────
all_results = []

def run_task(task_name, df_task, label_col, features, n_splits=5,
             cw=None, multiclass=False, xgb_scale_pos=None):
    print(f'\n{"="*60}')
    print(f'  {task_name}')
    print(f'{"="*60}')

    df_task = df_task[features + [label_col]].dropna(subset=[label_col]).copy()
    le      = LabelEncoder()
    y_full  = le.fit_transform(df_task[label_col])
    X_full  = df_task[features].values.astype(float)
    classes = le.classes_

    print(f'  N={len(y_full)}  |  {dict(zip(classes, np.bincount(y_full)))}')

    if len(np.unique(y_full)) < 2:
        print('  Skipping — only one class present.')
        return

    n_splits_eff = min(n_splits, int(np.bincount(y_full).min()))
    if n_splits_eff < 2:
        print('  Skipping — not enough samples for CV.')
        return
    cv = StratifiedKFold(n_splits=n_splits_eff, shuffle=True, random_state=42)

    # ── LR: restrict to high-coverage features + complete cases ──────────────
    lr_feats   = [f for f in features if coverage_pct.get(f, 0) >= MIN_COVERAGE]
    lr_idx     = [features.index(f) for f in lr_feats]
    X_lr_raw   = X_full[:, lr_idx]
    cc_mask    = ~np.isnan(X_lr_raw).any(axis=1)
    X_lr       = X_lr_raw[cc_mask]
    y_lr       = y_full[cc_mask]
    n_cc       = cc_mask.sum()
    print(f'  LR complete cases: {n_cc} / {len(y_full)}  '
          f'({len(lr_feats)} features ≥ {int(MIN_COVERAGE*100)}% coverage)')

    models = {
        'XGBoost':             (make_xgb(scale_pos=xgb_scale_pos if not multiclass else None),
                                X_full, y_full),
        'HistGradientBoosting':(make_hgb(cw=cw), X_full, y_full),
        'Logistic Regression': (make_lr(cw=cw),  X_lr,   y_lr),
    }

    best_auc, best_model_name, best_clf, best_X, best_y = -1, None, None, None, None

    for model_name, (clf, X_use, y_use) in models.items():
        if len(np.unique(y_use)) < 2:
            print(f'  {model_name}: skipping — only one class in subset')
            continue
        n_eff = min(n_splits_eff, int(np.bincount(y_use).min()))
        if n_eff < 2:
            print(f'  {model_name}: skipping — too few samples')
            continue
        cv_eff = StratifiedKFold(n_splits=n_eff, shuffle=True, random_state=42)
        try:
            y_prob = cross_val_predict(clf, X_use, y_use, cv=cv_eff, method='predict_proba')
            y_pred = y_prob.argmax(axis=1)

            cls_used = le.classes_  # same encoder for all
            if multiclass:
                auc = roc_auc_score(y_use, y_prob, multi_class='ovr', average='macro')
            else:
                auc = roc_auc_score(y_use, y_prob[:, 1])

            bal_acc = balanced_accuracy_score(y_use, y_pred)
            rpt     = classification_report(y_use, y_pred, target_names=cls_used,
                                            output_dict=True, zero_division=0)

            row = {
                'Task':           task_name,
                'Model':          model_name,
                'N (model)':      len(y_use),
                'N features':     len(lr_feats) if model_name == 'Logistic Regression' else len(features),
                'Missing strategy': ('Complete cases (≥80% cov features)'
                                     if model_name == 'Logistic Regression'
                                     else 'Native NaN'),
                'Classes':        ' / '.join(str(c) for c in cls_used),
                'ROC-AUC':        round(auc, 3),
                'Balanced Acc':   round(bal_acc, 3),
            }
            for cls in cls_used:
                row[f'Precision {cls}'] = round(rpt[cls]['precision'], 3)
                row[f'Recall {cls}']    = round(rpt[cls]['recall'],    3)
                row[f'F1 {cls}']        = round(rpt[cls]['f1-score'],  3)
            all_results.append(row)
            print(f'  {model_name:25s}  AUC={auc:.3f}  BalAcc={bal_acc:.3f}  N={len(y_use)}')

            if auc > best_auc:
                best_auc, best_model_name = auc, model_name
                best_clf, best_X, best_y  = clf, X_use, y_use

            # Confusion matrix
            if model_name == 'HistGradientBoosting':
                cm = confusion_matrix(y_use, y_pred)
                fig, ax = plt.subplots(figsize=(max(4, len(cls_used)), max(3, len(cls_used))))
                sns.heatmap(cm, annot=True, fmt='d', xticklabels=cls_used,
                            yticklabels=cls_used, cmap='Blues', ax=ax)
                ax.set_title(f'{task_name}\n{model_name} — Confusion Matrix (CV)')
                ax.set_xlabel('Predicted'); ax.set_ylabel('Actual')
                plt.tight_layout()
                save_fig(re.sub(r'[^\w]', '_', f'{task_name}_{model_name}') + '_cm.png')

        except Exception as e:
            print(f'  {model_name}: ERROR — {e}')

    # ── Feature importance — XGBoost (full fit, no imputation) ───────────────
    try:
        xgb_fi = make_xgb(scale_pos=xgb_scale_pos if not multiclass else None)
        xgb_fi.fit(X_full, y_full)
        fi = pd.Series(xgb_fi.feature_importances_, index=features).sort_values()
        fig, ax = plt.subplots(figsize=(6, max(4, len(features) * 0.32)))
        fi.plot.barh(ax=ax, color='steelblue')
        ax.set_title(f'{task_name}\nXGBoost Feature Importance (no imputation)')
        ax.axvline(fi.mean(), color='red', linestyle='--', linewidth=0.8, label='mean')
        ax.legend(fontsize=8)
        plt.tight_layout()
        save_fig(re.sub(r'[^\w]', '_', task_name) + '_feature_importance.png')
    except Exception as e:
        print(f'  Feature importance: {e}')

    # ── SHAP — XGBoost (best overall, or XGBoost if it ran) ──────────────────
    try:
        xgb_shap = make_xgb(scale_pos=xgb_scale_pos if not multiclass else None)
        xgb_shap.fit(X_full, y_full)
        explainer = shap.TreeExplainer(xgb_shap)
        shap_vals = explainer.shap_values(X_full)

        if isinstance(shap_vals, list):
            shap_arr = np.abs(np.array(shap_vals)).mean(axis=0)
        else:
            shap_arr = shap_vals

        plt.figure(figsize=(6, max(4, len(features) * 0.32)))
        shap.summary_plot(shap_arr, X_full, feature_names=features,
                          plot_type='bar', show=False)
        plt.title(f'{task_name}\nSHAP — XGBoost (no imputation)')
        plt.tight_layout()
        save_fig(re.sub(r'[^\w]', '_', task_name) + '_shap.png')
        plt.close()
    except Exception as e:
        print(f'  SHAP: {e}')


# ═══════════════════════════════════════════════════════════════════════════════
# TASK A — PD vs HC
# ═══════════════════════════════════════════════════════════════════════════════
df_a = df[df['label'].isin(['PD', 'HC'])].copy()
run_task('Task A — PD vs HC', df_a, 'label', FEAT_COLS, cw='balanced')

# ═══════════════════════════════════════════════════════════════════════════════
# TASK B — PD vs AP  (ET excluded)
# ═══════════════════════════════════════════════════════════════════════════════
df_b = df[df['det_code'].isin({0} | AP_CODES_NO_ET)].copy()
df_b['task_label'] = df_b['det_code'].apply(
    lambda c: 'PD' if c == 0 else ('AP' if c in AP_CODES_NO_ET else np.nan)
)
n_pd = (df_b['task_label'] == 'PD').sum()
n_ap = (df_b['task_label'] == 'AP').sum()
run_task('Task B — PD vs AP (ET excluded)', df_b, 'task_label', FEAT_COLS,
         cw='balanced', xgb_scale_pos=round(n_pd / max(n_ap, 1)))

# ═══════════════════════════════════════════════════════════════════════════════
# TASK C — PD vs HC vs AP  (3-class)
# ═══════════════════════════════════════════════════════════════════════════════
df_c = df[df['label'].isin(['PD', 'HC']) | df['det_code'].isin(AP_CODES)].copy()
df_c['task_label'] = df_c['det_code'].apply(
    lambda c: 'PD' if c == 0 else ('HC' if c == 10 else ('AP' if c in AP_CODES else np.nan))
)
run_task('Task C — PD vs HC vs AP', df_c, 'task_label', FEAT_COLS,
         cw='balanced', multiclass=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TASK D — AP subtype (exploratory)
# ═══════════════════════════════════════════════════════════════════════════════
df_d = df[df['det_code'].isin(AP_CODES)].copy()
df_d['task_label'] = df_d['det_code'].apply(
    lambda c: 'PSP' if c == 1 else ('MSA' if c == 2 else 'DLB+other')
)
run_task('Task D — AP subtype (exploratory)', df_d, 'task_label', FEAT_COLS,
         n_splits=3, cw='balanced', multiclass=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TASK E — HC vs Non-HC
# ═══════════════════════════════════════════════════════════════════════════════
df_e = df[df['label'].notna()].copy()
df_e['task_label'] = df_e['det_code'].apply(
    lambda c: 'HC' if c == 10 else ('Non-HC' if pd.notna(c) else np.nan)
)
n_hc     = (df_e['task_label'] == 'HC').sum()
n_non_hc = (df_e['task_label'] == 'Non-HC').sum()
run_task('Task E — HC vs Non-HC (PD+AP)', df_e, 'task_label', FEAT_COLS,
         cw='balanced', xgb_scale_pos=round(n_non_hc / max(n_hc, 1)))


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
print('\n\nWriting results...')
results_df = pd.DataFrame(all_results)

coverage_df = pd.DataFrame({
    'Feature':      FEAT_COLS,
    'Non-null N':   [int(df[c].notna().sum()) for c in FEAT_COLS],
    'Coverage %':   [round(coverage_pct[c] * 100, 1) for c in FEAT_COLS],
    'Used in LR':   ['Yes' if coverage_pct[c] >= MIN_COVERAGE else 'No' for c in FEAT_COLS],
    'Group': [
        'Clinical'          if c in ['disease_duration','age_at_onset','bmi','height_cm','weight_kg',
                                      'has_dyskinesia','has_freezing','has_falls','has_dementia',
                                      'has_dbs','has_motor_fluctuations','has_remission',
                                      'bilateral_onset','right_handed']
        else 'MoCA'         if c.startswith('moca')
        else 'UPDRS'        if c.startswith('updrs') or c.startswith('u3_')
        else 'SCOPA'        if c.startswith('scopa')
        else 'Demographic'  if c in ['age','edu_years','sex_male']
        else 'Neuropsychological'
        for c in FEAT_COLS
    ],
})

notes_df = pd.DataFrame([
    ['XGBoost',             'Native NaN', 'All features', 'NaN treated as a learnable direction at each split — no fill-in'],
    ['HistGradientBoosting','Native NaN', 'All features', 'sklearn gradient boosting with native NaN support'],
    ['Logistic Regression', f'Complete cases (≥{int(MIN_COVERAGE*100)}% cov features)',
     f'Features with ≥{int(MIN_COVERAGE*100)}% coverage only', 'Rows with any missing value dropped after feature filter'],
], columns=['Model', 'Missing Strategy', 'Features Used', 'Notes'])

task_notes = pd.DataFrame([
    ['Task A', 'PD vs HC',           'Binary',     f'{(df_a["label"]=="PD").sum()} PD / {(df_a["label"]=="HC").sum()} HC',  'class_weight=balanced'],
    ['Task B', 'PD vs AP',           'Binary',     f'{n_pd} PD / {n_ap} AP (ET excluded)',                                   'class_weight=balanced + XGB scale_pos_weight'],
    ['Task C', 'PD vs HC vs AP',     '3-class',    'PD / HC / AP combined',                                                  'class_weight=balanced'],
    ['Task D', 'AP subtype',         'Multiclass', 'PSP / MSA / DLB+other — EXPLORATORY',                                   '3-fold CV only'],
    ['Task E', 'HC vs Non-HC',       'Binary',     f'{n_hc} HC / {n_non_hc} Non-HC',                                        'class_weight=balanced + XGB scale_pos_weight'],
], columns=['Task', 'Description', 'Type', 'Classes/N', 'Notes'])

out_path = os.path.join(OUT_DIR, 'ml_results_v3.xlsx')
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils  import get_column_letter
import openpyxl

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    results_df.to_excel(writer,   sheet_name='Results',          index=False)
    coverage_df.to_excel(writer,  sheet_name='Feature Coverage', index=False)
    notes_df.to_excel(writer,     sheet_name='Model Notes',      index=False)
    task_notes.to_excel(writer,   sheet_name='Task Notes',       index=False)

wb  = openpyxl.load_workbook(out_path)
HDR = PatternFill('solid', fgColor='1F4E79')
HF  = Font(color='FFFFFF', bold=True)

TASK_FILLS = {
    'Task A': PatternFill('solid', fgColor='DFF0D8'),
    'Task B': PatternFill('solid', fgColor='FFF2CC'),
    'Task C': PatternFill('solid', fgColor='D9E1F2'),
    'Task D': PatternFill('solid', fgColor='FCE4D6'),
    'Task E': PatternFill('solid', fgColor='E2EFDA'),
}
LR_FILL  = PatternFill('solid', fgColor='FFF2CC')
YES_FILL = PatternFill('solid', fgColor='C6EFCE')
NO_FILL  = PatternFill('solid', fgColor='FFCCCC')

for sn in wb.sheetnames:
    ws = wb[sn]
    for cell in ws[1]:
        cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 50)

ws_r    = wb['Results']
task_col = next((i for i, c in enumerate(ws_r[1], 1) if c.value == 'Task'), None)
model_col = next((i for i, c in enumerate(ws_r[1], 1) if c.value == 'Model'), None)
if task_col:
    for row in ws_r.iter_rows(min_row=2):
        task_val  = str(row[task_col - 1].value or '')
        model_val = str(row[model_col - 1].value or '') if model_col else ''
        fill = next((f for k, f in TASK_FILLS.items() if k in task_val), None)
        if fill:
            for cell in row: cell.fill = fill
        if 'Logistic' in model_val:
            for cell in row: cell.fill = LR_FILL

ws_c = wb['Feature Coverage']
lr_col_idx = next((i for i, c in enumerate(ws_c[1], 1) if c.value == 'Used in LR'), None)
if lr_col_idx:
    for row in ws_c.iter_rows(min_row=2):
        val = str(row[lr_col_idx - 1].value or '')
        row[lr_col_idx - 1].fill = YES_FILL if val == 'Yes' else NO_FILL

wb.save(out_path)
print(f'  Results → {out_path}')
print(f'  Plots   → {PLOT_DIR}/')

print('\n── Summary ──────────────────────────────────────────────────────')
print(results_df[['Task', 'Model', 'N (model)', 'Missing strategy', 'ROC-AUC', 'Balanced Acc']]
      .to_string(index=False))
