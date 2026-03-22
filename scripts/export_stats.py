"""
export_stats.py
Builds cross_reference.xlsx with 4 sheets:
  - All Variables      — full data-dictionary × data cross-reference
  - Consolidated       — one row per base variable (all versions aggregated)
  - Numerical Stats    — descriptive stats per base variable (all versions pooled)
  - Categorical Stats  — value counts per base variable (all versions pooled)

Run:  python3 export_stats.py
"""

import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(ROOT, 'data')
DICT_PATH  = os.path.join(ROOT, 'reference', '2. COPN_DataDictionary_2025-09-24_annotated.xlsx')
EXCEL_PATH = os.path.join(ROOT, 'output', 'cross_reference.xlsx')


# ── Forms to drop entirely ────────────────────────────────────────────────────
EXCLUDE_FORMS = {'Apathy Evaluation Self', 'Apathy Evaluation Informant', 'FrSBe'}

# Columns that are always present and carry no substantive data
ADMIN_COLS = {'Project key', 'Event Name', 'Complete?'}


# ── Load datasets ─────────────────────────────────────────────────────────────
def load_dataset(name: str) -> pd.DataFrame:
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    # Drop rows where every non-admin column is null (skeleton rows)
    data_cols = [c for c in df.columns if c not in ADMIN_COLS]
    if data_cols:
        df = df[df[data_cols].notna().any(axis=1)].reset_index(drop=True)
    return df

datasets = {
    name: load_dataset(name)
    for name in os.listdir(DATA_DIR)
    if name not in EXCLUDE_FORMS                    # drop excluded forms entirely
    and not name.startswith('.')                    # skip hidden files
    and not os.path.splitext(name)[1] in            # skip files with known extensions
        {'.ipynb', '.csv', '.xlsx', '.py', '.txt'}
}
print(f'Loaded {len(datasets)} datasets (excluded: {sorted(EXCLUDE_FORMS)})')
for name, df in sorted(datasets.items()):
    print(f'  {name:40s}  {len(df):4d} rows')


# ── Load data dictionary ──────────────────────────────────────────────────────
data_dict = pd.read_excel(DICT_PATH)
print(f'Data dictionary: {len(data_dict)} variables')


# ── Form → file mapping ───────────────────────────────────────────────────────
FORM_TO_FILES = {
    'enrollment':                                    ['Enrollement'],
    'demographic_questionnaire':                     ['Demographic'],
    'clinical_questionnaire':                        ['Clinical'],
    'epidemiological_questionnaire':                 ['Epidemiological'],
    'mdsupdrs':                                      ['MDS-UPDRS'],
    'mdsupdrs_1':                                    ['MDS-UPDRS-1'],
    'moca':                                          ['MoCA'],
    'moca_1':                                        ['MoCA-1'],
    'moca_2':                                        ['MoCA-2'],
    'clinical_medications_questionnaire':            ['Medication'],
    'pdq39':                                         ['PDQ 39'],
    'pdq8':                                          ['PDQ 8'],
    'scopaauten':                                    ['SCOPA'],
    'bai':                                           ['BAI'],
    'ehibaibdi_ii_tests':                            ['BAI', 'BDII', 'EHI'],
    'ehi':                                           ['EHI'],
    'apathy_evaluation_scale_informant_qpn':         ['Apathy Evaluation Informant'],
    'apathy_evaluation_scale_selfrated_qpn':         ['Apathy Evaluation Self'],
    'apathy_scale':                                  ['Apathy Scale'],
    'fatigue_severity_scale':                        ['Fatigue Severity Scale'],
    'frsbe_formulaire_dautovaluation':               ['FrSBe'],
    'mbic_qpn':                                      ['MBIC'],
    'mild_behavioral_impairment_checklist_mibc':     ['MBIC (CaPRI)'],
    'neuropsychological_evaluation_qpn':             ['Neuropsychological'],
    'neuropsychological_evaluation_qpn_v02':         ['Neuropsychological V02'],
    'neuropsychological_evaluation_qpn_2':           ['Neuropsychological V02'],
    'neuropsychological_test_capri':                 ['Neuropsychological (CaPRI)'],
    'parkinson_anxiety_scale':                       ['Parkinson Severity Scale'],
    'schwab_and_england_activities_of_daily_living': ['Schwab & England'],
    'timed_up_go_test_tug':                          ['Timed Up Go'],
    'updrs':                                         ['UPDRS 1.2 part 3'],
}


# ── Column matching ───────────────────────────────────────────────────────────
def normalize_col(s) -> str:
    if pd.isna(s):
        return ''
    s = str(s).strip()
    s = re.sub(r'[\n\r\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()

file_col_norm = {
    fname: {normalize_col(c): c for c in df.columns}
    for fname, df in datasets.items()
}

def find_column(field_label, candidate_files):
    norm_label = normalize_col(field_label)
    if not norm_label or len(norm_label) < 5:
        return None, None
    for fname in candidate_files:
        if fname not in file_col_norm:
            continue
        col_index = file_col_norm[fname]
        if norm_label in col_index:
            return fname, col_index[norm_label]
        prefix = norm_label[:80]
        for nc, oc in col_index.items():
            if len(prefix) >= 15 and nc.startswith(prefix):
                return fname, oc
        for nc, oc in col_index.items():
            if len(nc) >= 15 and norm_label.startswith(nc[:80]):
                return fname, oc
    return None, None


# ── Build cross-reference (All Variables) ────────────────────────────────────
print('Building cross-reference...')
records = []
for _, row in data_dict.iterrows():
    var_name    = row['Variable / Field Name']
    form_name   = str(row['Form Name']) if pd.notna(row['Form Name']) else ''
    field_label = row['Field Label']
    role        = row['Role']
    impl_type   = row['Implementation type']
    choices     = row['Choices, Calculations, OR Slider Labels']

    candidate_files = FORM_TO_FILES.get(form_name, [])
    matched_file, matched_col = find_column(field_label, candidate_files)

    missing_pct, n_unique, sample_values = None, None, ''
    if matched_file and matched_col:
        s = datasets[matched_file][matched_col]
        missing_pct  = round(s.isna().mean() * 100, 1)
        n_unique     = int(s.nunique(dropna=True))
        top_vals     = s.dropna().value_counts().head(3).index.tolist()
        sample_values = ' | '.join(str(v)[:25] for v in top_vals)

    label_short   = normalize_col(field_label)[:90] if pd.notna(field_label) else ''
    choices_short = str(choices)[:120] if pd.notna(choices) else ''
    primary_file  = matched_file or (candidate_files[0] if candidate_files else 'UNMAPPED')

    records.append({
        'Variable Name':   var_name,
        'File':            primary_file,
        'Role':            role,
        'Impl. Level':     str(impl_type) if pd.notna(impl_type) else '',
        'Found in Data':   'Yes' if matched_col else 'No',
        'Field Label':     label_short,
        'Choices / Formula': choices_short,
        '% Missing':       missing_pct,
        'N Unique Values': n_unique,
        'Top Values':      sample_values,
        '_matched_file':   matched_file or '',
        '_matched_col':    matched_col  or '',
    })

cross_ref = pd.DataFrame(records)
found = (cross_ref['Found in Data'] == 'Yes').sum()
print(f'  Matched {found}/{len(cross_ref)} variables ({found/len(cross_ref)*100:.1f}%)')


# ── Build Consolidated (base variable grouping) ───────────────────────────────
all_var_names = set(cross_ref['Variable Name'])

def get_base_name(var):
    m = re.match(r'^(.+?)_(v\d+|v0\d+)$', str(var))
    if m:
        return m.group(1)
    m = re.match(r'^(.+?)_(\d+)$', str(var))
    if m and m.group(1) in all_var_names:
        return m.group(1)
    return var

cross_ref['base_name']    = cross_ref['Variable Name'].apply(get_base_name)
cross_ref['is_versioned'] = cross_ref['base_name'] != cross_ref['Variable Name']

# ── Ungroup numeric-suffix versions that are genuinely different variables ─────
# If all versions of a base name live in the same file AND have different field
# labels, they represent distinct timepoints/contexts — keep them separate.
for base, g in cross_ref[cross_ref['is_versioned']].groupby('base_name'):
    # Explicit _vN/_v0N suffixes are always true versions — leave grouped
    if g['Variable Name'].str.match(r'^.+_(v\d+|v0\d+)$').all():
        continue
    # Numeric suffix (_2, _3, …): ungroup if same file but different labels
    if g['File'].nunique() == 1:
        n_distinct_labels = g['Field Label'].dropna().apply(normalize_col).nunique()
        if n_distinct_labels > 1:
            cross_ref.loc[g.index, 'base_name']    = cross_ref.loc[g.index, 'Variable Name']
            cross_ref.loc[g.index, 'is_versioned'] = False

def aggregate_group(base, g):
    found_mask   = g['Found in Data'] == 'Yes'
    missing_vals = g.loc[found_mask, '% Missing'].dropna()
    base_row     = g[g['Variable Name'] == base]
    ref_row      = base_row.iloc[0] if len(base_row) else g.iloc[0]
    combined_top = ' | '.join(dict.fromkeys(
        v.strip() for cell in g['Top Values'].dropna()
        for v in str(cell).split(' | ') if v.strip()
    ))[:200]
    return {
        'Base Variable':        base,
        'All Versions':         ' | '.join(sorted(g['Variable Name'].unique())),
        'N Versions':           g['Variable Name'].nunique(),
        'Files':                ' | '.join(sorted(set(g['File'].dropna()))),
        'Role':                 ref_row['Role'],
        'Impl. Level':          ref_row['Impl. Level'],
        'Field Label':          ref_row['Field Label'],
        'Choices / Formula':    ref_row['Choices / Formula'],
        'Found in Any Version': 'Yes' if found_mask.any() else 'No',
        'Found in All Versions':'Yes' if found_mask.all() else 'No',
        'N Versions Found':     int(found_mask.sum()),
        'Min % Missing':        round(missing_vals.min(), 1) if len(missing_vals) else None,
        'Avg % Missing':        round(missing_vals.mean(), 1) if len(missing_vals) else None,
        'Max % Missing':        round(missing_vals.max(), 1) if len(missing_vals) else None,
        'Combined Top Values':  combined_top,
    }

versioned_agg = pd.DataFrame([
    aggregate_group(base, g)
    for base, g in cross_ref[cross_ref['is_versioned']].groupby('base_name', sort=False)
])

non_versioned = cross_ref[~cross_ref['is_versioned']].copy()
non_versioned_agg = non_versioned.rename(columns={'Variable Name': 'Base Variable'}).assign(
    **{'All Versions':          non_versioned['Variable Name'],
       'N Versions':            1,
       'Files':                 non_versioned['File'],
       'Found in Any Version':  non_versioned['Found in Data'],
       'Found in All Versions': non_versioned['Found in Data'],
       'N Versions Found':      (non_versioned['Found in Data'] == 'Yes').astype(int),
       'Min % Missing':         non_versioned['% Missing'],
       'Avg % Missing':         non_versioned['% Missing'],
       'Max % Missing':         non_versioned['% Missing'],
       'Combined Top Values':   non_versioned['Top Values']}
)[list(versioned_agg.columns)]

consolidated = pd.concat([versioned_agg, non_versioned_agg], ignore_index=True)
print(f'Consolidated: {len(consolidated)} base variables')


# ── Days-since-dx helpers (defined here; used before and after rename) ─────────
_clin_raw = load_dataset('Clinical')
_dx_col   = [c for c in _clin_raw.columns if 'date of diagnosis' in c.lower()][0]
dx_lookup = pd.to_datetime(
    _clin_raw[['Project key', _dx_col]].dropna()
    .drop_duplicates('Project key')
    .set_index('Project key')[_dx_col],
    errors='coerce'
)   # Series: Project key → diagnosis date

DATE_PATTERNS = [
    'questionnaire completed', 'questionnaire rempli',
    'tests completed',         'tests complétés',
    'assessment completed',    'évaluation remplie',
    'date of moca administration',
    'neuropsycholgical test date',
    'date of study visit',
]

def _find_completion_date_col(df: pd.DataFrame):
    """First column matching a date-header pattern AND containing YYYY-MM-DD values."""
    for pat in DATE_PATTERNS:
        for col in df.columns:
            if pat in col.lower():
                sample = df[col].dropna().head(5)
                if sample.apply(lambda v: bool(re.match(r'\d{4}-\d{2}', str(v)))).any():
                    return col
    return None

# Capture completion-date column names BEFORE renaming ──────────────────────────
completion_date_cols: dict[str, str] = {}    # {ds_name: raw_col_name}
for ds_name, df in datasets.items():
    col = _find_completion_date_col(df)
    if col is not None:
        completion_date_cols[ds_name] = col


# ── Rename dataset columns → dictionary variable names ────────────────────────
# Build per-file rename maps from the matched cross_ref rows.
# If two variables matched the same column, first match wins (stable iteration order).
rename_maps: dict[str, dict[str, str]] = {}   # {fname: {original_col: var_name}}
for _, row in cross_ref.iterrows():
    if row['Found in Data'] != 'Yes' or not row['_matched_col']:
        continue
    fname = row['_matched_file']
    col   = row['_matched_col']
    var   = row['Variable Name']
    rename_maps.setdefault(fname, {})
    if col not in rename_maps[fname]:          # first match wins
        rename_maps[fname][col] = var

conflicts = 0
for fname, rmap in rename_maps.items():
    datasets[fname] = datasets[fname].rename(columns=rmap)
    conflicts += sum(1 for oc, vn in rmap.items() if oc == vn)  # no-op renames

renamed_total = sum(len(r) for r in rename_maps.values())
print(f'Renamed {renamed_total} columns across {len(rename_maps)} datasets')

# ── Second pass: rename versioned variable names → base variable names ─────────
# Only rename a column if its base name would be unique within that dataset
# (avoids duplicate column names when multiple versions exist in the same file).
base_renamed_total = 0
for fname, df in datasets.items():
    base_map: dict[str, str] = {}
    base_targets: dict[str, int] = {}
    for col in df.columns:
        base = get_base_name(col)
        base_targets[base] = base_targets.get(base, 0) + 1
    for col in df.columns:
        base = get_base_name(col)
        if base != col and base_targets[base] == 1:
            base_map[col] = base
    if base_map:
        datasets[fname] = df.rename(columns=base_map)
        base_renamed_total += len(base_map)
print(f'Base-variable pass: renamed {base_renamed_total} columns across {len(datasets)} datasets')


# ── Insert 'Days since dx' after renaming ─────────────────────────────────────
# Formula: [form completion date] – [Clinical dx date]  (integer days)
added = 0
for ds_name, raw_col in completion_date_cols.items():
    df = datasets[ds_name]
    if 'Project key' not in df.columns:
        continue
    # raw_col may have been renamed; look up the new name
    renamed_col = rename_maps.get(ds_name, {}).get(raw_col, raw_col)
    if renamed_col not in df.columns:
        continue
    completion = pd.to_datetime(df[renamed_col], errors='coerce')
    dx         = df['Project key'].map(dx_lookup)
    days       = (completion - dx).dt.days
    anchor     = 'Event Name' if 'Event Name' in df.columns else 'Project key'
    df.insert(df.columns.get_loc(anchor) + 1, 'Days since dx', days)
    datasets[ds_name] = df
    added += 1

print(f'Added "Days since dx" to {added} datasets')


# ── Build stats by pooling all versions per base variable ─────────────────────
print('Computing stats from pooled versions...')

# Map version name → (file, renamed_col).
# After renaming, the column in the dataset IS the variable name.
version_to_data: dict[str, tuple[str, str]] = {}
for _, row in cross_ref.iterrows():
    if row['Found in Data'] != 'Yes' or not row['_matched_col']:
        continue
    fname        = row['_matched_file']
    var          = row['Variable Name']
    original_col = row['_matched_col']
    renamed_col  = rename_maps.get(fname, {}).get(original_col, original_col)
    # If the base-variable pass renamed this column further, use the new name
    base_col = get_base_name(renamed_col)
    if base_col != renamed_col and fname in datasets and base_col in datasets[fname].columns:
        renamed_col = base_col
    version_to_data[var] = (fname, renamed_col)

def pool_series(all_versions_str: str) -> pd.Series:
    """Concatenate data from all matched versions of a base variable."""
    parts = []
    for ver in re.split(r'\s*\|\s*', str(all_versions_str)):
        ver = ver.strip()
        if ver in version_to_data:
            fname, col = version_to_data[ver]
            parts.append(datasets[fname][col].dropna())
    return pd.concat(parts) if parts else pd.Series(dtype=object)

def col_dtype(series: pd.Series) -> str:
    n_unique = series.nunique()
    if pd.api.types.is_numeric_dtype(series) and n_unique > 12:
        return 'Numerical'
    if n_unique <= 30 or pd.api.types.is_object_dtype(series):
        return 'Categorical'
    return 'Text/Free'

def num_stats(series: pd.Series) -> dict:
    s = pd.to_numeric(series, errors='coerce').dropna()
    if len(s) == 0:
        return {}
    return {
        'Count':  int(len(s)),
        'Mean':   round(float(s.mean()),         3),
        'Median': round(float(s.median()),       3),
        'Std':    round(float(s.std()),          3),
        'Min':    round(float(s.min()),          3),
        'Q1':     round(float(s.quantile(0.25)), 3),
        'Q3':     round(float(s.quantile(0.75)), 3),
        'Max':    round(float(s.max()),          3),
        'Skew':   round(float(s.skew()),         3),
        'Kurt':   round(float(s.kurtosis()),     3),
    }

def cat_stats(series: pd.Series) -> list[dict]:
    vc = series.value_counts(dropna=True)
    total = vc.sum()
    return [
        {'Value':   str(v).split('/')[0].strip(),
         'Count':   int(c),
         'Pct (%)': round(100 * c / total, 1)}
        for v, c in vc.items()
    ]

num_rows: list[dict] = []
cat_rows: list[dict] = []

for _, row in consolidated.iterrows():
    if row['Found in Any Version'] != 'Yes':
        continue

    base      = row['Base Variable']
    role      = row['Role']
    impl      = row['Impl. Level']
    label     = row['Field Label']
    versions  = row['All Versions']
    n_vers    = row['N Versions Found']
    files     = row['Files']

    pooled = pool_series(versions)
    if len(pooled) == 0:
        continue

    total_obs   = len(pooled)
    missing_pct = round(float(row['Avg % Missing']) if pd.notna(row['Avg % Missing']) else 0, 1)
    dtype       = col_dtype(pooled)

    base_info = {
        'Base Variable': base,
        'File':          files,
        'Role':          role,
        'Impl. Level':   impl,
        'Field Label':   str(label)[:80],
        'N Versions':    n_vers,
        'Non-null':      total_obs,
        'Avg Missing %': missing_pct,
    }

    if dtype == 'Numerical':
        stats = num_stats(pooled)
        if stats:
            num_rows.append({**base_info, **stats})

    elif dtype == 'Categorical':
        for cat_row in cat_stats(pooled):
            cat_rows.append({**base_info, **cat_row})

num_df = pd.DataFrame(num_rows)
cat_df = pd.DataFrame(cat_rows)
print(f'  Numerical Stats : {len(num_df):,} rows')
print(f'  Categorical Stats: {len(cat_df):,} rows')


# ── Export table for All Variables (drop internal columns) ────────────────────
table = cross_ref.drop(columns=['is_versioned', '_matched_file', '_matched_col'])
table.insert(1, 'Base Variable', table.pop('base_name'))


# ═══════════════════════════════════════════════════════════════════════════════
# DIAGNOSIS CONSISTENCY FLAGS
# Fields checked:
#   [Enrollment]  Enrolment Group
#   [Clinical]    Determined diagnosis  (0=PD,1=PSP,2=MSA,3=CBS,4=DLB,6=ET,7=RBD)
#   [Clinical]    Was the patient diagnosed with Parkinson's disease? (Yes/No/Uncertain)
#   [Clinical]    1a. If No/Uncertain, is the diagnosis... (alternative dx)
#   [Clinical]    If other, please specify (free-text)
# ═══════════════════════════════════════════════════════════════════════════════
print('Building diagnosis flags...')

# Load raw (pre-rename) versions so column-name lookups are reliable
_enroll = load_dataset('Enrollement')
_clin   = load_dataset('Clinical')

ENROL_COL = [c for c in _enroll.columns if 'nrolment' in c and 'roup' in c][0]
DIAG_COL  = [c for c in _clin.columns   if 'Determined diagnosis' in c][0]
PD_COL    = [c for c in _clin.columns   if 'Was the patient diagnosed' in c][0]
ALT_COL   = [c for c in _clin.columns   if c.strip().startswith("1a.")][0]
SPEC_COL  = [c for c in _clin.columns   if 'please specify' in c.lower()
                                          and 'Veuillez' in c
                                          and 'autre' in c.lower()
                                          and '2' not in c[-5:]  ][0]

# Numeric code → short diagnosis label
DIAG_LABEL = {0:'PD', 1:'PSP', 2:'MSA', 3:'CBS', 4:'DLB', 6:'ET', 7:'RBD'}

# 1a text → expected Determined diagnosis code
ALT_TO_CODE = {
    'Progressive Supranuclear Palsy (PSP)': 1,
    'Multiple System Atrophy (MSA)':        2,
    'Corticobasal Syndrome (CBS)':          3,
    'Dementia with Lewy Bodies (DLB)':      4,
    'Essential Tremor (ET)':                6,
    'REM Sleep Behaviour Disorder (RBD)':   7,
}

def _short_enrol(val):
    if pd.isna(val): return None
    s = str(val).split('/')[0].strip()
    if 'PD' in s or "Parkinson's Disease" in s: return 'PD'
    if 'AP' in s or 'Atypical' in s:            return 'AP'
    if 'Healthy' in s:                          return 'HC'
    return s

def _short_pd(val):
    if pd.isna(val): return None
    return str(val).split('/')[0].strip()   # Yes / No / Uncertain

def _alt_code(val):
    """Extract expected Determined code from 1a free-text value."""
    if pd.isna(val): return None
    for key, code in ALT_TO_CODE.items():
        if key.split(' (')[0] in str(val):
            return code
    return None  # Not Determined / Other

# Build one merged row per participant
_e = _enroll[['Project key', ENROL_COL]].copy()
_e['enrol_group'] = _e[ENROL_COL].apply(_short_enrol)

_c = _clin[['Project key', DIAG_COL, PD_COL, ALT_COL, SPEC_COL]].copy()
_c['det_code']  = pd.to_numeric(_c[DIAG_COL], errors='coerce')
_c['det_label'] = _c['det_code'].map(DIAG_LABEL)
_c['pd_yn']     = _c[PD_COL].apply(_short_pd)
_c['alt_code']  = _c[ALT_COL].apply(_alt_code)
_c['alt_text']  = _c[ALT_COL].str.split('/').str[0].str.replace(r'^\.*\s*', '', regex=True).str.strip()
_c['spec_text'] = _c[SPEC_COL]

merged = _e.merge(_c, on='Project key', how='outer')

flag_rows   = []
review_rows = []   # all participants — flagged or not
for _, r in merged.iterrows():
    pid        = r['Project key']
    enrol      = r['enrol_group']
    det_code   = r['det_code']
    det_label  = r['det_label']
    pd_yn      = r['pd_yn']
    alt_code   = r['alt_code']
    alt_text   = r['alt_text']
    spec_text  = r['spec_text']

    flags_for_row = []

    # ── Flag A: Enrolment Group vs Determined Diagnosis ───────────────────────
    if enrol == 'PD' and pd.notna(det_code) and det_code != 0:
        flags_for_row.append(
            f'Enrolled as PD but Determined diagnosis = {det_label} ({int(det_code)})'
        )
    if enrol == 'AP' and pd.notna(det_code) and det_code == 0:
        flags_for_row.append(
            'Enrolled as AP but Determined diagnosis = PD (0)'
        )

    # ── Flag B: Was diagnosed with PD? vs Determined Diagnosis ───────────────
    if pd_yn == 'Yes' and pd.notna(det_code) and det_code != 0:
        flags_for_row.append(
            f'Was diagnosed = Yes but Determined diagnosis = {det_label} ({int(det_code)})'
        )
    if pd_yn in ('No', 'Uncertain') and pd.notna(det_code) and det_code == 0:
        flags_for_row.append(
            f'Was diagnosed = {pd_yn} but Determined diagnosis = PD (0)'
        )

    # ── Flag C: Enrolment Group vs Was diagnosed with PD? ────────────────────
    if enrol == 'PD' and pd_yn in ('No', 'Uncertain'):
        flags_for_row.append(
            f'Enrolled as PD but "Was diagnosed with PD?" = {pd_yn}'
        )
    if enrol == 'AP' and pd_yn == 'Yes':
        flags_for_row.append(
            'Enrolled as AP but "Was diagnosed with PD?" = Yes'
        )

    # ── Flag D: 1a alternative diagnosis vs Determined Diagnosis ─────────────
    if alt_code is not None and pd.notna(det_code) and alt_code != det_code:
        flags_for_row.append(
            f'1a alternative dx ({alt_text}) expects code {alt_code} '
            f'but Determined = {det_label} ({int(det_code)})'
        )
    # 1a is filled but Determined is NaN
    if pd.notna(r[ALT_COL]) and pd.isna(det_code):
        flags_for_row.append(
            f'1a filled ({alt_text}) but Determined diagnosis is missing'
        )
    # Determined is non-PD but 1a is empty (and Was_PD is No/Uncertain)
    if pd.notna(det_code) and det_code != 0 and pd.isna(r[ALT_COL]) and pd_yn in ('No','Uncertain'):
        flags_for_row.append(
            f'Determined = {det_label} ({int(det_code)}) but 1a is missing'
        )

    # ── Build not-flagged reason ──────────────────────────────────────────────
    if not flags_for_row:
        if pd.isna(enrol) and pd.isna(det_code):
            not_flagged = 'No enrollment or clinical data'
        elif pd.isna(enrol):
            not_flagged = 'No enrollment record — cannot check against clinical'
        elif pd.isna(det_code):
            not_flagged = f'{enrol} enrolled but no Determined Dx in clinical data'
        elif enrol == 'PD' and det_code == 0:
            not_flagged = 'PD enrolled, Determined = PD (0): consistent'
        elif enrol == 'AP' and det_code != 0:
            not_flagged = f'AP enrolled, Determined = {det_label} ({int(det_code)}): consistent'
        elif enrol == 'HC':
            not_flagged = 'HC enrolled: diagnosis flags not applicable'
        else:
            not_flagged = 'No inconsistencies detected'
    else:
        not_flagged = ''

    base_row = {
        'Project key':          pid,
        'Enrolment Group':      r[ENROL_COL] if pd.notna(r[ENROL_COL]) else '',
        'Determined Dx (code)': int(det_code) if pd.notna(det_code) else '',
        'Determined Dx':        det_label or '',
        'Was Dx with PD?':      pd_yn or '',
        '1a Alternative Dx':    alt_text if pd.notna(r[ALT_COL]) else '',
        'Other (specify)':      str(spec_text) if pd.notna(spec_text) else '',
        'Status':               'Flagged' if flags_for_row else 'OK',
        'N Flags':              len(flags_for_row),
        'Flags':                ' | '.join(flags_for_row),
        'Not Flagged Reason':   not_flagged,
    }
    review_rows.append(base_row)

    if flags_for_row:
        flag_rows.append({k: v for k, v in base_row.items()
                          if k not in ('Status', 'Not Flagged Reason')})

diag_flags_df = pd.DataFrame(flag_rows).sort_values(
    ['N Flags', 'Project key'], ascending=[False, True]
)
diag_review_df = pd.DataFrame(review_rows).sort_values(
    ['N Flags', 'Project key'], ascending=[False, True]
)
_n_total = merged['Project key'].nunique()
print(f'  Flagged participants: {len(diag_flags_df):,} ({len(diag_flags_df)/_n_total*100:.1f}%)')
print(f'  Diagnosis Review total: {len(diag_review_df):,} participants')


# ── Sheet name helper (Excel max = 31 chars, no special chars) ────────────────
def safe_sheet_name(name: str) -> str:
    clean = re.sub(r'[\\/*?\[\]:]', '', name)
    return clean[:31]


# ── Completeness Matrix ────────────────────────────────────────────────────────
print('Building completeness matrix...')

ADMIN_FOR_COMPLETENESS = {'Project key', 'Event Name', 'Complete?', 'Days since dx'}

# Per-participant, per-form % missing
comp_records = []
for ds_name, df in sorted(datasets.items()):
    data_cols = [c for c in df.columns if c not in ADMIN_FOR_COMPLETENESS]
    if not data_cols or 'Project key' not in df.columns:
        continue
    n_data = len(data_cols)
    for _, row in df.iterrows():
        pid = row['Project key']
        n_missing = int(row[data_cols].isna().sum())
        pct_missing = round(100 * n_missing / n_data, 1)
        comp_records.append({'Project key': pid, 'Form': ds_name, 'Pct Missing': pct_missing})

comp_long = pd.DataFrame(comp_records)

# Pivot: participants × forms (average if multiple visits per form)
comp_matrix = comp_long.pivot_table(
    index='Project key', columns='Form', values='Pct Missing', aggfunc='mean'
).reset_index()
comp_matrix.columns.name = None

# Round averaged values
form_cols_matrix = [c for c in comp_matrix.columns if c != 'Project key']
comp_matrix[form_cols_matrix] = comp_matrix[form_cols_matrix].round(1)

# Attach Enrolment Group label
enrol_lookup = _e.drop_duplicates('Project key').set_index('Project key')['enrol_group']
comp_matrix.insert(1, 'Enrolment Group', comp_matrix['Project key'].map(enrol_lookup))

# Sort: group, then participant
comp_matrix = comp_matrix.sort_values(
    ['Enrolment Group', 'Project key'], na_position='last'
).reset_index(drop=True)

print(f'  Completeness Matrix: {len(comp_matrix):,} participants × {len(form_cols_matrix)} forms')


# ── Enrollment Coverage ────────────────────────────────────────────────────────
# For every project key in Enrollment: which forms do they appear in (have data)?
# Participants missing from a form may simply not have data there.
print('Building enrollment coverage...')

# Keys present in each form (after skeleton-row removal)
_form_keys: dict[str, set] = {
    ds: set(df['Project key'].dropna().unique())
    for ds, df in datasets.items()
    if 'Project key' in df.columns and ds != 'Enrollement'
}
_coverage_forms = sorted(_form_keys.keys())

# Full enrollment list with group label (_e has enrol_group added by _short_enrol)
_enroll_all = _e[['Project key', ENROL_COL, 'enrol_group']].drop_duplicates('Project key')

coverage_rows = []
for _, er in _enroll_all.iterrows():
    pid   = er['Project key']
    group = er['enrol_group'] or ''
    row: dict = {
        'Project key':    pid,
        'Enrolment Group': er[ENROL_COL] if pd.notna(er[ENROL_COL]) else '',
        'Enrolment Status': group,
    }
    n_present = n_missing = 0
    for form in _coverage_forms:
        present = pid in _form_keys[form]
        row[form] = 'Yes' if present else 'No'
        if present: n_present += 1
        else:       n_missing += 1
    row['N Forms Present'] = n_present
    row['N Forms Missing'] = n_missing
    coverage_rows.append(row)

coverage_df = (
    pd.DataFrame(coverage_rows)
    .sort_values(['Enrolment Status', 'N Forms Missing', 'Project key'],
                 ascending=[True, False, True])
    .reset_index(drop=True)
)
print(f'  Enrollment Coverage: {len(coverage_df):,} participants × {len(_coverage_forms)} forms')


# ── Completeness by Diagnosis Summary ─────────────────────────────────────────
summary_rows = []
groups = comp_matrix.groupby('Enrolment Group', dropna=False)
for group_label, group_df in groups:
    label = group_label if pd.notna(group_label) else 'Unknown / Missing'
    for form in form_cols_matrix:
        vals = group_df[form].dropna()
        if len(vals) == 0:
            continue
        summary_rows.append({
            'Enrolment Group': label,
            'Form':            form,
            'N Entries':       int(len(vals)),
            'Avg % Missing':   round(float(vals.mean()), 1),
        })

comp_by_diag = pd.DataFrame(summary_rows)
print(f'  Completeness by Diagnosis: {len(comp_by_diag):,} rows')


# ── Write all sheets in one pass ─────────────────────────────────────────────
print(f'Writing {EXCEL_PATH}...')
with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
    # Reference sheets
    table.to_excel(writer,           sheet_name='All Variables',         index=False)
    consolidated.to_excel(writer,    sheet_name='Consolidated',          index=False)
    num_df.to_excel(writer,          sheet_name='Numerical Stats',       index=False)
    cat_df.to_excel(writer,          sheet_name='Categorical Stats',     index=False)
    diag_flags_df.to_excel(writer,   sheet_name='Diagnosis Flags',       index=False)
    diag_review_df.to_excel(writer,  sheet_name='Diagnosis Review',      index=False)
    comp_matrix.to_excel(writer,     sheet_name='Completeness Matrix',   index=False)
    comp_by_diag.to_excel(writer,    sheet_name='Completeness by Dx',    index=False)
    coverage_df.to_excel(writer,     sheet_name='Enrollment Coverage',   index=False)
    # One sheet per cleaned dataset (skeleton rows removed, columns renamed)
    for ds_name, df in sorted(datasets.items()):
        df.to_excel(writer, sheet_name=safe_sheet_name(ds_name), index=False)
    print(f'  Added {len(datasets)} dataset sheets')


# ── Formatting ────────────────────────────────────────────────────────────────
ROLE_FILL = {
    'Predictor':         PatternFill('solid', fgColor='D0E8FF'),
    'Outcome':           PatternFill('solid', fgColor='FFD0D0'),
    'Outcome/Diagnosis': PatternFill('solid', fgColor='FFE8CC'),
    'Admin':             PatternFill('solid', fgColor='F0F0F0'),
}
IMPL_FILL = {
    'Level 1': PatternFill('solid', fgColor='C6EFCE'),
    'Level 2': PatternFill('solid', fgColor='BDD7EE'),
    'Level 3': PatternFill('solid', fgColor='FFEB9C'),
    'Level 4': PatternFill('solid', fgColor='FFC7CE'),
}
HDR_FILL   = PatternFill('solid', fgColor='212529')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=11)
GREEN_FONT = Font(color='155724', bold=True)
RED_FONT   = Font(color='721C24', italic=True)
EVEN_FILL  = PatternFill('solid', fgColor='F8F9FA')
ODD_FILL   = PatternFill('solid', fgColor='FFFFFF')

def miss_fill(v: float) -> PatternFill:
    if v > 70: return PatternFill('solid', fgColor='FFC7CE')
    if v > 40: return PatternFill('solid', fgColor='FFEB9C')
    return         PatternFill('solid', fgColor='C6EFCE')

def autowidth(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 60)

def style_dict_sheet(ws, role_col, impl_col, found_col, missing_col):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=2):
        role_val  = row[role_col - 1].value
        impl_val  = row[impl_col - 1].value
        found_val = row[found_col - 1].value
        miss_val  = row[missing_col - 1].value
        for cell in row:
            if role_val in ROLE_FILL:
                cell.fill = ROLE_FILL[role_val]
        if impl_val in IMPL_FILL:
            row[impl_col - 1].fill = IMPL_FILL[impl_val]
        if found_val == 'Yes':  row[found_col - 1].font = GREEN_FONT
        elif found_val == 'No': row[found_col - 1].font = RED_FONT
        try:
            row[missing_col - 1].fill = miss_fill(float(miss_val))
        except (ValueError, TypeError):
            pass
    autowidth(ws)

def style_stats_sheet(ws, missing_col_idx: int):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
        try:
            row[missing_col_idx - 1].fill = miss_fill(
                float(row[missing_col_idx - 1].value or 0))
        except (ValueError, TypeError):
            pass
    autowidth(ws)

def style_data_sheet(ws):
    """Freeze header, style header row, alternate row shading, auto-width."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
    autowidth(ws)

FLAG_RED    = PatternFill('solid', fgColor='FFC7CE')
FLAG_ORANGE = PatternFill('solid', fgColor='FFEB9C')

def style_flags_sheet(ws):
    """Header + alternate rows; highlight rows with multiple flags in red."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    # N Flags is column 8 (index 7)
    n_flags_col = 8
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        try:
            n = int(row[n_flags_col - 1].value or 0)
        except (ValueError, TypeError):
            n = 0
        rf = FLAG_RED if n > 1 else (FLAG_ORANGE if n == 1 else (EVEN_FILL if i % 2 == 0 else ODD_FILL))
        for cell in row:
            cell.fill = rf
    autowidth(ws)

def style_completeness_matrix(ws):
    """Header row + heatmap coloring: green=low missing, red=high missing."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'C2'   # freeze Project key + Enrolment Group
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None or cell.column <= 2:
                continue
            try:
                cell.fill = miss_fill(float(cell.value))
            except (ValueError, TypeError):
                pass
    autowidth(ws)

def style_completeness_summary(ws):
    """Header + zebra rows; Avg % Missing column gets heatmap."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    # Avg % Missing is column 4
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
        try:
            row[3].fill = miss_fill(float(row[3].value or 0))
        except (ValueError, TypeError):
            pass
    autowidth(ws)

wb = openpyxl.load_workbook(EXCEL_PATH)
# All Variables: Role=3, Impl=4, Found=5, Missing=8
style_dict_sheet(wb['All Variables'],  role_col=3, impl_col=4, found_col=5, missing_col=8)
# Consolidated: Role=5, Impl=6, Found=9, Missing=12
style_dict_sheet(wb['Consolidated'],   role_col=5, impl_col=6, found_col=9, missing_col=12)
# Stats sheets: Avg Missing % is col 8
style_stats_sheet(wb['Numerical Stats'],   missing_col_idx=8)
style_stats_sheet(wb['Categorical Stats'], missing_col_idx=8)
# Diagnosis flags sheet
style_flags_sheet(wb['Diagnosis Flags'])

def style_review_sheet(ws):
    """Flagged rows red/orange at top; OK rows green; header dark."""
    OK_FILL = PatternFill('solid', fgColor='C6EFCE')
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    status_col = next(
        (i + 1 for i, cell in enumerate(ws[1]) if cell.value == 'Status'), None
    )
    n_flags_col = next(
        (i + 1 for i, cell in enumerate(ws[1]) if cell.value == 'N Flags'), None
    )
    for row in ws.iter_rows(min_row=2):
        status = row[status_col - 1].value if status_col else ''
        try:
            n = int(row[n_flags_col - 1].value or 0) if n_flags_col else 0
        except (ValueError, TypeError):
            n = 0
        if status == 'Flagged':
            rf = FLAG_RED if n > 1 else FLAG_ORANGE
        else:
            rf = OK_FILL
        for cell in row:
            cell.fill = rf
    autowidth(ws)

style_review_sheet(wb['Diagnosis Review'])
# Completeness sheets
style_completeness_matrix(wb['Completeness Matrix'])
style_completeness_summary(wb['Completeness by Dx'])

def style_coverage_sheet(ws):
    """Header dark; Yes = green, No = red; freeze first 3 cols."""
    YES_FILL = PatternFill('solid', fgColor='C6EFCE')
    NO_FILL  = PatternFill('solid', fgColor='FFC7CE')
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'D2'   # freeze Project key + 2 group cols
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == 'Yes':
                cell.fill = YES_FILL
            elif cell.value == 'No':
                cell.fill = NO_FILL
            else:
                cell.fill = EVEN_FILL
    autowidth(ws)

style_coverage_sheet(wb['Enrollment Coverage'])
# Dataset sheets
for ds_name in sorted(datasets.keys()):
    style_data_sheet(wb[safe_sheet_name(ds_name)])
wb.save(EXCEL_PATH)

print(f'\nDone. Sheets in {EXCEL_PATH}:')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row - 1:,} rows × {ws.max_column} cols')
