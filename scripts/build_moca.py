"""
build_moca.py
Combines MoCA, MoCA-1, MoCA-2 into a single MoCA_combined form.
For participants with two or more forms completed, calculates:
  - Time between forms (days)
  - Difference in scores for each numerical variable

Outputs: moca_analysis.xlsx
  Sheet 1: MoCA_combined  — all entries with Source Form column
  Sheet 2: MoCA_delta     — one row per consecutive pair for multi-form participants

Run: python3 build_moca.py
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, 'data')
OUT_PATH = os.path.join(ROOT, 'output', 'moca_analysis.xlsx')

ADMIN_COLS = {'Project key', 'Event Name', 'Complete?', 'Source Form'}

DATE_COL = 'Date of MoCA administration     Date d\'administration du MoCA'

# Numeric score columns used for delta calculations (subscores + total)
SCORE_COLS = [
    'Visuospatial/Executive Score    Score Visuospatial/Éxécutif ',
    'Naming Score    Score Dénomination',
    'Attention Score    Score Attention',
    'Language Score    Score Langage',
    'Abstraction Score    Score Abstraction',
    'Delayed Recall Score    Score Rappel',
    'Orientation Score    Score Orientation',
    'TOTAL SCORE (make sure to include extra point for 12 years or less of education):'
    '    SCORE TOTAL (assurez-vous d\'inclure un point supplémentaire pour 12 ans ou '
    'moins d\'éducation) : ',
]

# ── Load and clean each MoCA form ─────────────────────────────────────────────
print('Loading MoCA forms...')

def load_moca(name: str) -> pd.DataFrame:
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    data_cols = [c for c in df.columns if c not in {'Project key', 'Event Name', 'Complete?'}]
    df = df[df[data_cols].notna().any(axis=1)].reset_index(drop=True)
    df.insert(0, 'Source Form', name)
    return df

dfs = [load_moca(name) for name in ['MoCA', 'MoCA-1', 'MoCA-2']]
for df, name in zip(dfs, ['MoCA', 'MoCA-1', 'MoCA-2']):
    print(f'  {name}: {len(df)} rows with data')

# ── Combine ────────────────────────────────────────────────────────────────────
moca_combined = pd.concat(dfs, ignore_index=True)
moca_combined[DATE_COL] = pd.to_datetime(moca_combined[DATE_COL], errors='coerce')

# Resolve the actual TOTAL SCORE column name (may be truncated in the variable above)
total_col = next((c for c in moca_combined.columns if c.startswith('TOTAL SCORE')), None)
if total_col and total_col not in SCORE_COLS:
    SCORE_COLS[-1] = total_col   # fix the name in case of whitespace difference

# Keep only score cols that actually exist
SCORE_COLS = [c for c in SCORE_COLS if c in moca_combined.columns]

print(f'\nCombined: {len(moca_combined)} rows, '
      f'{moca_combined["Project key"].nunique()} unique participants')
print(f'Score columns for delta: {len(SCORE_COLS)}')

# ── Find multi-form participants ───────────────────────────────────────────────
entry_counts = moca_combined.groupby('Project key').size()
multi_pids   = entry_counts[entry_counts >= 2].index
print(f'Participants with 2+ forms: {len(multi_pids)}')

from collections import Counter
combos = Counter(
    tuple(sorted(
        moca_combined[moca_combined['Project key'] == pid]['Source Form'].tolist()
    ))
    for pid in multi_pids
)
for combo, cnt in sorted(combos.items(), key=lambda x: -x[1]):
    print(f'  {" + ".join(combo)}: {cnt} participants')

# ── Compute consecutive-pair deltas ───────────────────────────────────────────
print('\nComputing deltas...')

delta_rows = []
for pid in multi_pids:
    grp = (moca_combined[moca_combined['Project key'] == pid]
           .sort_values(DATE_COL, na_position='last')
           .reset_index(drop=True))

    for i in range(len(grp) - 1):
        r1, r2 = grp.iloc[i], grp.iloc[i + 1]
        d1, d2 = r1[DATE_COL], r2[DATE_COL]
        days = int((d2 - d1).days) if (pd.notna(d1) and pd.notna(d2)) else np.nan

        row: dict = {
            'Project key':   pid,
            'Source Form 1': r1['Source Form'],
            'Source Form 2': r2['Source Form'],
            'Date 1':        d1.date() if pd.notna(d1) else np.nan,
            'Date 2':        d2.date() if pd.notna(d2) else np.nan,
            'Days Between':  days,
        }

        for col in SCORE_COLS:
            v1 = pd.to_numeric(r1.get(col), errors='coerce')
            v2 = pd.to_numeric(r2.get(col), errors='coerce')
            short = col.split('    ')[0].strip()   # use English half as header
            if pd.notna(v1) or pd.notna(v2):
                row[f'Δ {short}'] = (v2 - v1) if (pd.notna(v1) and pd.notna(v2)) else np.nan
            row[f'{short} (form 1)'] = v1
            row[f'{short} (form 2)'] = v2

        delta_rows.append(row)

delta_df = pd.DataFrame(delta_rows)
print(f'Delta rows: {len(delta_df)} consecutive pairs from {len(multi_pids)} participants')

# ── Formatting helpers ─────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='212529')
HDR_FONT  = Font(color='FFFFFF', bold=True, size=11)
EVEN_FILL = PatternFill('solid', fgColor='F8F9FA')
ODD_FILL  = PatternFill('solid', fgColor='FFFFFF')
POS_FILL  = PatternFill('solid', fgColor='C6EFCE')   # green = improved (Δ < 0... wait)
NEG_FILL  = PatternFill('solid', fgColor='FFC7CE')   # red = declined

SOURCE_FILLS = {
    'MoCA':   PatternFill('solid', fgColor='D0E8FF'),
    'MoCA-1': PatternFill('solid', fgColor='BDD7EE'),
    'MoCA-2': PatternFill('solid', fgColor='C8E6C9'),
}

def autowidth(ws, max_w=55):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)

def style_combined(ws):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        src = row[0].value or ''   # Source Form is col 1
        rf  = SOURCE_FILLS.get(src, EVEN_FILL if i % 2 == 0 else ODD_FILL)
        for cell in row:
            cell.fill = rf
    autowidth(ws)

def style_delta(ws):
    headers = [c.value for c in ws[1]]
    delta_start = next((i for i, h in enumerate(headers) if isinstance(h, str) and h.startswith('Δ ')), None)

    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.freeze_panes = 'A2'

    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
        if delta_start is not None:
            for cell in row[delta_start:]:
                h = headers[cell.column - 1]
                if isinstance(h, str) and h.startswith('Δ '):
                    try:
                        v = float(cell.value)
                        # Higher MoCA score = better, so Δ > 0 is improvement
                        if v > 0:
                            cell.fill = POS_FILL
                        elif v < 0:
                            cell.fill = NEG_FILL
                    except (ValueError, TypeError):
                        pass
    autowidth(ws)

# ── Write output ──────────────────────────────────────────────────────────────
print(f'\nWriting {OUT_PATH}...')

# Sort combined: multi-form participants first (for easy review), then singles
moca_combined['_is_multi'] = moca_combined['Project key'].isin(multi_pids)
moca_combined = (moca_combined
                 .sort_values(['_is_multi', 'Project key', DATE_COL],
                              ascending=[False, True, True])
                 .drop(columns=['_is_multi'])
                 .reset_index(drop=True))

with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    moca_combined.to_excel(writer, sheet_name='MoCA_combined', index=False)
    delta_df.to_excel(writer,      sheet_name='MoCA_delta',    index=False)

wb = openpyxl.load_workbook(OUT_PATH)
style_combined(wb['MoCA_combined'])
style_delta(wb['MoCA_delta'])
wb.save(OUT_PATH)

print(f'\nDone. Sheets in {OUT_PATH}:')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row - 1:,} rows × {ws.max_column} cols')
