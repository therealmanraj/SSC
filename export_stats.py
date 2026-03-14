"""
export_stats.py
Builds cross_reference.xlsx with 4 sheets:
  - All Variables      — full data-dictionary × data cross-reference
  - Consolidated       — one row per base variable (all versions aggregated)
  - Numerical Stats    — descriptive stats per base variable (all versions pooled)
  - Categorical Stats  — value counts per base variable (all versions pooled)

Run:  python3 export_stats.py
"""

import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT       = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = os.path.join(ROOT, 'data')
DICT_PATH  = os.path.join(ROOT, 'Help Files', '2. COPN_DataDictionary_2025-09-24_annotated.xlsx')
EXCEL_PATH = os.path.join(ROOT, 'cross_reference.xlsx')


# ── Forms to drop entirely ────────────────────────────────────────────────────
EXCLUDE_FORMS = {'Apathy Evaluation Self', 'Apathy Evaluation Informant', 'FrSBe'}

# Columns that are always present and carry no substantive data
ADMIN_COLS = {'Project key', 'Event Name', 'Complete?'}


# ── Load datasets ─────────────────────────────────────────────────────────────
def load_dataset(name: str) -> pd.DataFrame:
    df = pd.read_csv(os.path.join(DATA_DIR, name))
    if df.columns[0].startswith('Unnamed'):
        df = df.iloc[:, 1:]
    # Drop rows where every non-admin column is null (skeleton rows)
    data_cols = [c for c in df.columns if c not in ADMIN_COLS]
    if data_cols:
        df = df[df[data_cols].notna().any(axis=1)].reset_index(drop=True)
    return df

datasets = {
    name: load_dataset(name)
    for name in os.listdir(DATA_DIR)
    if name not in EXCLUDE_FORMS          # drop excluded forms entirely
}
print(f'Loaded {len(datasets)} datasets (excluded: {sorted(EXCLUDE_FORMS)})')
for name, df in sorted(datasets.items()):
    print(f'  {name:40s}  {len(df):4d} rows')


# ── Load data dictionary ──────────────────────────────────────────────────────
data_dict = pd.read_excel(DICT_PATH)
print(f'Data dictionary: {len(data_dict)} variables')


# ── Form → file mapping ───────────────────────────────────────────────────────
FORM_TO_FILES = {
    'enrollment':                                    ['Enrollement'],
    'demographic_questionnaire':                     ['Demographic'],
    'clinical_questionnaire':                        ['Clinical'],
    'epidemiological_questionnaire':                 ['Epidemiological'],
    'mdsupdrs':                                      ['MDS-UPDRS'],
    'mdsupdrs_1':                                    ['MDS-UPDRS-1'],
    'moca':                                          ['MoCA'],
    'moca_1':                                        ['MoCA-1'],
    'moca_2':                                        ['MoCA-2'],
    'clinical_medications_questionnaire':            ['Medication'],
    'pdq39':                                         ['PDQ 39'],
    'pdq8':                                          ['PDQ 8'],
    'scopaauten':                                    ['SCOPA'],
    'bai':                                           ['BAI'],
    'ehibaibdi_ii_tests':                            ['BAI', 'BDII', 'EHI'],
    'ehi':                                           ['EHI'],
    'apathy_evaluation_scale_informant_qpn':         ['Apathy Evaluation Informant'],
    'apathy_evaluation_scale_selfrated_qpn':         ['Apathy Evaluation Self'],
    'apathy_scale':                                  ['Apathy Scale'],
    'fatigue_severity_scale':                        ['Fatigue Severity Scale'],
    'frsbe_formulaire_dautovaluation':               ['FrSBe'],
    'mbic_qpn':                                      ['MBIC'],
    'mild_behavioral_impairment_checklist_mibc':     ['MBIC (CaPRI)'],
    'neuropsychological_evaluation_qpn':             ['Neuropsychological'],
    'neuropsychological_evaluation_qpn_v02':         ['Neuropsychological V02'],
    'neuropsychological_evaluation_qpn_2':           ['Neuropsychological V02'],
    'neuropsychological_test_capri':                 ['Neuropsychological (CaPRI)'],
    'parkinson_anxiety_scale':                       ['Parkinson Severity Scale'],
    'schwab_and_england_activities_of_daily_living': ['Schwab & England'],
    'timed_up_go_test_tug':                          ['Timed Up Go'],
    'updrs':                                         ['UPDRS 1.2 part 3'],
}


# ── Column matching ───────────────────────────────────────────────────────────
def normalize_col(s) -> str:
    if pd.isna(s):
        return ''
    s = str(s).strip()
    s = re.sub(r'[\n\r\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()

file_col_norm = {
    fname: {normalize_col(c): c for c in df.columns}
    for fname, df in datasets.items()
}

def find_column(field_label, candidate_files):
    norm_label = normalize_col(field_label)
    if not norm_label or len(norm_label) < 5:
        return None, None
    for fname in candidate_files:
        if fname not in file_col_norm:
            continue
        col_index = file_col_norm[fname]
        if norm_label in col_index:
            return fname, col_index[norm_label]
        prefix = norm_label[:80]
        for nc, oc in col_index.items():
            if len(prefix) >= 15 and nc.startswith(prefix):
                return fname, oc
        for nc, oc in col_index.items():
            if len(nc) >= 15 and norm_label.startswith(nc[:80]):
                return fname, oc
    return None, None


# ── Build cross-reference (All Variables) ────────────────────────────────────
print('Building cross-reference...')
records = []
for _, row in data_dict.iterrows():
    var_name    = row['Variable / Field Name']
    form_name   = str(row['Form Name']) if pd.notna(row['Form Name']) else ''
    field_label = row['Field Label']
    role        = row['Role']
    impl_type   = row['Implementation type']
    choices     = row['Choices, Calculations, OR Slider Labels']

    candidate_files = FORM_TO_FILES.get(form_name, [])
    matched_file, matched_col = find_column(field_label, candidate_files)

    missing_pct, n_unique, sample_values = None, None, ''
    if matched_file and matched_col:
        s = datasets[matched_file][matched_col]
        missing_pct  = round(s.isna().mean() * 100, 1)
        n_unique     = int(s.nunique(dropna=True))
        top_vals     = s.dropna().value_counts().head(3).index.tolist()
        sample_values = ' | '.join(str(v)[:25] for v in top_vals)

    label_short   = normalize_col(field_label)[:90] if pd.notna(field_label) else ''
    choices_short = str(choices)[:120] if pd.notna(choices) else ''
    primary_file  = matched_file or (candidate_files[0] if candidate_files else 'UNMAPPED')

    records.append({
        'Variable Name':   var_name,
        'File':            primary_file,
        'Role':            role,
        'Impl. Level':     str(impl_type) if pd.notna(impl_type) else '',
        'Found in Data':   'Yes' if matched_col else 'No',
        'Field Label':     label_short,
        'Choices / Formula': choices_short,
        '% Missing':       missing_pct,
        'N Unique Values': n_unique,
        'Top Values':      sample_values,
        '_matched_file':   matched_file or '',
        '_matched_col':    matched_col  or '',
    })

cross_ref = pd.DataFrame(records)
found = (cross_ref['Found in Data'] == 'Yes').sum()
print(f'  Matched {found}/{len(cross_ref)} variables ({found/len(cross_ref)*100:.1f}%)')


# ── Build Consolidated (base variable grouping) ───────────────────────────────
all_var_names = set(cross_ref['Variable Name'])

def get_base_name(var):
    m = re.match(r'^(.+?)_(v\d+|v0\d+)$', str(var))
    if m:
        return m.group(1)
    m = re.match(r'^(.+?)_(\d+)$', str(var))
    if m and m.group(1) in all_var_names:
        return m.group(1)
    return var

cross_ref['base_name']    = cross_ref['Variable Name'].apply(get_base_name)
cross_ref['is_versioned'] = cross_ref['base_name'] != cross_ref['Variable Name']

def aggregate_group(base, g):
    found_mask   = g['Found in Data'] == 'Yes'
    missing_vals = g.loc[found_mask, '% Missing'].dropna()
    base_row     = g[g['Variable Name'] == base]
    ref_row      = base_row.iloc[0] if len(base_row) else g.iloc[0]
    combined_top = ' | '.join(dict.fromkeys(
        v.strip() for cell in g['Top Values'].dropna()
        for v in str(cell).split(' | ') if v.strip()
    ))[:200]
    return {
        'Base Variable':        base,
        'All Versions':         ' | '.join(sorted(g['Variable Name'].unique())),
        'N Versions':           g['Variable Name'].nunique(),
        'Files':                ' | '.join(sorted(set(g['File'].dropna()))),
        'Role':                 ref_row['Role'],
        'Impl. Level':          ref_row['Impl. Level'],
        'Field Label':          ref_row['Field Label'],
        'Choices / Formula':    ref_row['Choices / Formula'],
        'Found in Any Version': 'Yes' if found_mask.any() else 'No',
        'Found in All Versions':'Yes' if found_mask.all() else 'No',
        'N Versions Found':     int(found_mask.sum()),
        'Min % Missing':        round(missing_vals.min(), 1) if len(missing_vals) else None,
        'Avg % Missing':        round(missing_vals.mean(), 1) if len(missing_vals) else None,
        'Max % Missing':        round(missing_vals.max(), 1) if len(missing_vals) else None,
        'Combined Top Values':  combined_top,
    }

versioned_agg = pd.DataFrame([
    aggregate_group(base, g)
    for base, g in cross_ref[cross_ref['is_versioned']].groupby('base_name', sort=False)
])

non_versioned = cross_ref[~cross_ref['is_versioned']].copy()
non_versioned_agg = non_versioned.rename(columns={'Variable Name': 'Base Variable'}).assign(
    **{'All Versions':          non_versioned['Variable Name'],
       'N Versions':            1,
       'Files':                 non_versioned['File'],
       'Found in Any Version':  non_versioned['Found in Data'],
       'Found in All Versions': non_versioned['Found in Data'],
       'N Versions Found':      (non_versioned['Found in Data'] == 'Yes').astype(int),
       'Min % Missing':         non_versioned['% Missing'],
       'Avg % Missing':         non_versioned['% Missing'],
       'Max % Missing':         non_versioned['% Missing'],
       'Combined Top Values':   non_versioned['Top Values']}
)[list(versioned_agg.columns)]

consolidated = pd.concat([versioned_agg, non_versioned_agg], ignore_index=True)
print(f'Consolidated: {len(consolidated)} base variables')


# ── Rename dataset columns → dictionary variable names ────────────────────────
# Build per-file rename maps from the matched cross_ref rows.
# If two variables matched the same column, first match wins (stable iteration order).
rename_maps: dict[str, dict[str, str]] = {}   # {fname: {original_col: var_name}}
for _, row in cross_ref.iterrows():
    if row['Found in Data'] != 'Yes' or not row['_matched_col']:
        continue
    fname = row['_matched_file']
    col   = row['_matched_col']
    var   = row['Variable Name']
    rename_maps.setdefault(fname, {})
    if col not in rename_maps[fname]:          # first match wins
        rename_maps[fname][col] = var

conflicts = 0
for fname, rmap in rename_maps.items():
    datasets[fname] = datasets[fname].rename(columns=rmap)
    conflicts += sum(1 for oc, vn in rmap.items() if oc == vn)  # no-op renames

renamed_total = sum(len(r) for r in rename_maps.values())
print(f'Renamed {renamed_total} columns across {len(rename_maps)} datasets')


# ── Build stats by pooling all versions per base variable ─────────────────────
print('Computing stats from pooled versions...')

# Map version name → (file, renamed_col).
# After renaming, the column in the dataset IS the variable name.
version_to_data: dict[str, tuple[str, str]] = {}
for _, row in cross_ref.iterrows():
    if row['Found in Data'] != 'Yes' or not row['_matched_col']:
        continue
    fname        = row['_matched_file']
    var          = row['Variable Name']
    original_col = row['_matched_col']
    renamed_col  = rename_maps.get(fname, {}).get(original_col, original_col)
    version_to_data[var] = (fname, renamed_col)

def pool_series(all_versions_str: str) -> pd.Series:
    """Concatenate data from all matched versions of a base variable."""
    parts = []
    for ver in re.split(r'\s*\|\s*', str(all_versions_str)):
        ver = ver.strip()
        if ver in version_to_data:
            fname, col = version_to_data[ver]
            parts.append(datasets[fname][col].dropna())
    return pd.concat(parts) if parts else pd.Series(dtype=object)

def col_dtype(series: pd.Series) -> str:
    n_unique = series.nunique()
    if pd.api.types.is_numeric_dtype(series) and n_unique > 12:
        return 'Numerical'
    if n_unique <= 30 or pd.api.types.is_object_dtype(series):
        return 'Categorical'
    return 'Text/Free'

def num_stats(series: pd.Series) -> dict:
    s = pd.to_numeric(series, errors='coerce').dropna()
    if len(s) == 0:
        return {}
    return {
        'Count':  int(len(s)),
        'Mean':   round(float(s.mean()),         3),
        'Median': round(float(s.median()),       3),
        'Std':    round(float(s.std()),          3),
        'Min':    round(float(s.min()),          3),
        'Q1':     round(float(s.quantile(0.25)), 3),
        'Q3':     round(float(s.quantile(0.75)), 3),
        'Max':    round(float(s.max()),          3),
        'Skew':   round(float(s.skew()),         3),
        'Kurt':   round(float(s.kurtosis()),     3),
    }

def cat_stats(series: pd.Series) -> list[dict]:
    vc = series.value_counts(dropna=True)
    total = vc.sum()
    return [
        {'Value':   str(v).split('/')[0].strip(),
         'Count':   int(c),
         'Pct (%)': round(100 * c / total, 1)}
        for v, c in vc.items()
    ]

num_rows: list[dict] = []
cat_rows: list[dict] = []

for _, row in consolidated.iterrows():
    if row['Found in Any Version'] != 'Yes':
        continue

    base      = row['Base Variable']
    role      = row['Role']
    impl      = row['Impl. Level']
    label     = row['Field Label']
    versions  = row['All Versions']
    n_vers    = row['N Versions Found']
    files     = row['Files']

    pooled = pool_series(versions)
    if len(pooled) == 0:
        continue

    total_obs   = len(pooled)
    missing_pct = round(float(row['Avg % Missing']) if pd.notna(row['Avg % Missing']) else 0, 1)
    dtype       = col_dtype(pooled)

    base_info = {
        'Base Variable': base,
        'File':          files,
        'Role':          role,
        'Impl. Level':   impl,
        'Field Label':   str(label)[:80],
        'N Versions':    n_vers,
        'Non-null':      total_obs,
        'Avg Missing %': missing_pct,
    }

    if dtype == 'Numerical':
        stats = num_stats(pooled)
        if stats:
            num_rows.append({**base_info, **stats})

    elif dtype == 'Categorical':
        for cat_row in cat_stats(pooled):
            cat_rows.append({**base_info, **cat_row})

num_df = pd.DataFrame(num_rows)
cat_df = pd.DataFrame(cat_rows)
print(f'  Numerical Stats : {len(num_df):,} rows')
print(f'  Categorical Stats: {len(cat_df):,} rows')


# ── Export table for All Variables (drop internal columns) ────────────────────
table = cross_ref.drop(columns=['base_name', 'is_versioned', '_matched_file', '_matched_col'])


# ── Sheet name helper (Excel max = 31 chars, no special chars) ────────────────
def safe_sheet_name(name: str) -> str:
    clean = re.sub(r'[\\/*?\[\]:]', '', name)
    return clean[:31]


# ── Write all sheets in one pass ─────────────────────────────────────────────
print(f'Writing {EXCEL_PATH}...')
with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
    # Reference sheets
    table.to_excel(writer,        sheet_name='All Variables',    index=False)
    consolidated.to_excel(writer, sheet_name='Consolidated',     index=False)
    num_df.to_excel(writer,       sheet_name='Numerical Stats',  index=False)
    cat_df.to_excel(writer,       sheet_name='Categorical Stats',index=False)
    # One sheet per cleaned dataset (skeleton rows removed, columns renamed)
    for ds_name, df in sorted(datasets.items()):
        df.to_excel(writer, sheet_name=safe_sheet_name(ds_name), index=False)
    print(f'  Added {len(datasets)} dataset sheets')


# ── Formatting ────────────────────────────────────────────────────────────────
ROLE_FILL = {
    'Predictor':         PatternFill('solid', fgColor='D0E8FF'),
    'Outcome':           PatternFill('solid', fgColor='FFD0D0'),
    'Outcome/Diagnosis': PatternFill('solid', fgColor='FFE8CC'),
    'Admin':             PatternFill('solid', fgColor='F0F0F0'),
}
IMPL_FILL = {
    'Level 1': PatternFill('solid', fgColor='C6EFCE'),
    'Level 2': PatternFill('solid', fgColor='BDD7EE'),
    'Level 3': PatternFill('solid', fgColor='FFEB9C'),
    'Level 4': PatternFill('solid', fgColor='FFC7CE'),
}
HDR_FILL   = PatternFill('solid', fgColor='212529')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=11)
GREEN_FONT = Font(color='155724', bold=True)
RED_FONT   = Font(color='721C24', italic=True)
EVEN_FILL  = PatternFill('solid', fgColor='F8F9FA')
ODD_FILL   = PatternFill('solid', fgColor='FFFFFF')

def miss_fill(v: float) -> PatternFill:
    if v > 70: return PatternFill('solid', fgColor='FFC7CE')
    if v > 40: return PatternFill('solid', fgColor='FFEB9C')
    return         PatternFill('solid', fgColor='C6EFCE')

def autowidth(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 60)

def style_dict_sheet(ws, role_col, impl_col, found_col, missing_col):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=2):
        role_val  = row[role_col - 1].value
        impl_val  = row[impl_col - 1].value
        found_val = row[found_col - 1].value
        miss_val  = row[missing_col - 1].value
        for cell in row:
            if role_val in ROLE_FILL:
                cell.fill = ROLE_FILL[role_val]
        if impl_val in IMPL_FILL:
            row[impl_col - 1].fill = IMPL_FILL[impl_val]
        if found_val == 'Yes':  row[found_col - 1].font = GREEN_FONT
        elif found_val == 'No': row[found_col - 1].font = RED_FONT
        try:
            row[missing_col - 1].fill = miss_fill(float(miss_val))
        except (ValueError, TypeError):
            pass
    autowidth(ws)

def style_stats_sheet(ws, missing_col_idx: int):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
        try:
            row[missing_col_idx - 1].fill = miss_fill(
                float(row[missing_col_idx - 1].value or 0))
        except (ValueError, TypeError):
            pass
    autowidth(ws)

def style_data_sheet(ws):
    """Freeze header, style header row, alternate row shading, auto-width."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        rf = EVEN_FILL if i % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill = rf
    autowidth(ws)

wb = openpyxl.load_workbook(EXCEL_PATH)
# All Variables: Role=3, Impl=4, Found=5, Missing=8
style_dict_sheet(wb['All Variables'],  role_col=3, impl_col=4, found_col=5, missing_col=8)
# Consolidated: Role=5, Impl=6, Found=9, Missing=12
style_dict_sheet(wb['Consolidated'],   role_col=5, impl_col=6, found_col=9, missing_col=12)
# Stats sheets: Avg Missing % is col 8
style_stats_sheet(wb['Numerical Stats'],   missing_col_idx=8)
style_stats_sheet(wb['Categorical Stats'], missing_col_idx=8)
# Dataset sheets
for ds_name in sorted(datasets.keys()):
    style_data_sheet(wb[safe_sheet_name(ds_name)])
wb.save(EXCEL_PATH)

print(f'\nDone. Sheets in {EXCEL_PATH}:')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row - 1:,} rows × {ws.max_column} cols')
