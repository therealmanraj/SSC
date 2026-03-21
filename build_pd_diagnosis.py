"""
build_pd_diagnosis.py
Implements the "Steps to derive dx" logic from PD_Diagnosis.xlsx.

Algorithm (from Steps to derive dx sheet):
  1. IF Determined Dx (D) not blank          → Derived Dx = D
  2. ELSE IF G not blank
         IF G = "Other"                      → Derived Dx = H (free-text)
         ELSE                                → Derived Dx = G
  3. ELSE (G blank)                          → Step 2b: fall back to Enrolment Group
         B = PD  → "PD"  (proxy, low confidence)
         B = HC  → "HC"  (proxy)
         B = AP  → "AP"  (proxy, subtype unknown)
         B = blank → null (no information)

Note: E="Yes/Oui" and blank D never co-occur in this dataset (confirmed).
      Therefore the E="Yes/Oui" branch in the algorithm is never triggered.

Certainty mapping (from Col I):
  >90%           → High
  50-89%         → Moderate
  <50%           → Low
  Not applicable → Not applicable (treat as high confidence — confirmed cases)
  Unknown        → Unknown
  Blank          → Unknown
  Proxy (Step2b) → Proxy (low confidence)

Outputs: pd_diagnosis.xlsx with sheets:
  PD_Diagnosis       — one row per participant with derived dx + certainty
  Step2b_proxy       — participants where Enrolment Group was used as fallback
  Free_text_dx       — participants with free-text "Other (specify)" dx needing review
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT     = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.path.join(ROOT, 'New Data to do', 'PD_Diagnosis.xlsx')
OUT_PATH = os.path.join(ROOT, 'pd_diagnosis.xlsx')

# ── Load data ──────────────────────────────────────────────────────────────────
data = pd.read_excel(SRC_PATH, sheet_name='Data')
c = list(data.columns)

COL_B = c[1]   # Enrolment Group
COL_D = c[3]   # Determined Dx (group name)
COL_E = c[4]   # Was Dx with PD?
COL_F = c[5]   # Alternative Dx (text)
COL_G = c[6]   # Alternative Dx (group name)
COL_H = c[7]   # Other (specify)
COL_I = c[8]   # Dx level of certainty

print(f'Loaded {len(data)} rows from Data sheet')

# ── Certainty normalisation ────────────────────────────────────────────────────
CERTAINTY_MAP = {
    '>90% (high certainty/haute certitude)': 'High (>90%)',
    '50-89% (likely/probable)':              'Moderate (50-89%)',
    '< 50% (unlikely/peu probable)':         'Low (<50%)',
    'Not applicable/Sans objet':             'Not applicable (high confidence)',
    'Unknown/inconnu':                       'Unknown',
}

def map_certainty(val, is_proxy: bool = False) -> str:
    if is_proxy:
        return 'Proxy (low confidence)'
    if pd.isna(val) or str(val).strip() == '':
        return 'Unknown'
    return CERTAINTY_MAP.get(str(val).strip(), str(val).strip())


# ── Derivation logic ───────────────────────────────────────────────────────────
ENROLMENT_PD = "PD (Parkinson's Disease)/(Maladie de Parkinson)"
ENROLMENT_HC = 'Healthy control/Contrôle'
ENROLMENT_AP = 'AP (Atypical Parkinsonism)/(Parkinsonisme Atypique)'

rows_out = []

for _, row in data.iterrows():
    D = row[COL_D]
    E = row[COL_E]
    F = row[COL_F]
    G = row[COL_G]
    H = row[COL_H]
    B = row[COL_B]
    I = row[COL_I]
    pid = row[c[0]]

    derived_dx   = None
    source       = None
    is_proxy     = False
    needs_review = False  # free-text H entries

    # Step 1 — Determined Dx is filled
    if pd.notna(D) and str(D).strip():
        derived_dx = str(D).strip()
        source     = 'Determined Dx (Col D)'

    # Step 2 / 3 — Determined Dx blank; check Alternative Dx
    elif pd.notna(G) and str(G).strip():
        g_val = str(G).strip()
        if g_val == 'Other':
            if pd.notna(H) and str(H).strip():
                derived_dx   = str(H).strip()
                source       = 'Other (specify) (Col H)'
                needs_review = True
            else:
                # G = "Other" but H blank → no label recoverable → Step 2b
                pass
        else:
            derived_dx = g_val
            source     = 'Alternative Dx (Col G)'

    # Step 2b — Last resort: fall back to Enrolment Group
    # Applies when:
    #   (a) E is blank — form was never completed, OR
    #   (b) E = No/Uncertain but G is blank — form completed but no dx recoverable
    #       (per narrative Step 3: "No label can be recovered → go to Step 2b")
    if derived_dx is None:
        e_val = str(E).strip() if pd.notna(E) else ''
        b_val = str(B).strip() if pd.notna(B) else ''
        src_qualifier = 'form not completed' if not e_val else 'form completed, no dx recoverable'

        if b_val == ENROLMENT_PD:
            derived_dx = 'PD'
            source     = f'Enrolment Group proxy — {src_qualifier} (Col B)'
            is_proxy   = True
        elif b_val == ENROLMENT_HC:
            derived_dx = 'HC'
            source     = f'Enrolment Group proxy — {src_qualifier} (Col B)'
            is_proxy   = True
        elif b_val == ENROLMENT_AP:
            derived_dx = 'AP'
            source     = f'Enrolment Group proxy — {src_qualifier} (Col B)'
            is_proxy   = True
        else:
            # B blank — no information at all
            derived_dx = 'Not Determined'
            source     = 'No information available'

    certainty = map_certainty(I, is_proxy=is_proxy)

    rows_out.append({
        'Project key':        pid,
        'Enrolment Group':    B if pd.notna(B) else '',
        'Derived Dx':         derived_dx if derived_dx else '',
        'Dx Source':          source,
        'Is Proxy':           is_proxy,
        'Needs Review':       needs_review,
        'Certainty (raw)':    str(I).strip() if pd.notna(I) else '',
        'Certainty':          certainty,
        'Determined Dx':      str(D).strip() if pd.notna(D) else '',
        'Was Dx with PD?':    str(E).strip() if pd.notna(E) else '',
        'Alt Dx (group)':     str(G).strip() if pd.notna(G) else '',
        'Alt Dx (text)':      str(F).strip() if pd.notna(F) else '',
        'Other (specify)':    str(H).strip() if pd.notna(H) else '',
    })

result = pd.DataFrame(rows_out)

# ── Summary ────────────────────────────────────────────────────────────────────
print('\n=== Derived Dx distribution ===')
print(result['Derived Dx'].replace('', 'null/blank').value_counts(dropna=False).to_string())

print('\n=== Dx Source distribution ===')
print(result['Dx Source'].value_counts(dropna=False).to_string())

print('\n=== Certainty distribution ===')
print(result['Certainty'].value_counts(dropna=False).to_string())

print(f'\nProxy rows:        {result["Is Proxy"].sum()}')
print(f'Free-text (review):{result["Needs Review"].sum()}')
print(f'No dx recoverable: {(result["Derived Dx"]=="").sum()}')

# ── Sub-tables ─────────────────────────────────────────────────────────────────
proxy_df  = result[result['Is Proxy']].copy()
review_df = result[result['Needs Review']].copy()

# ── Verify against existing Formula - Derived Dx ──────────────────────────────
# Map our output back to compare with "Formula - Derived Dx"
formula_col = c[14]
merged = data[['Project key', formula_col, c[13], c[15]]].merge(
    result[['Project key', 'Derived Dx', 'Dx Source']],
    on='Project key', how='left'
)
# Normalise formula col for comparison (strip enrolment group verbosity)
def normalise(v):
    if pd.isna(v): return ''
    s = str(v).strip()
    if 'Parkinson' in s and 'PD' in s: return 'PD'
    if 'Healthy' in s: return 'HC'
    if 'Atypical' in s or 'Parkinsonisme' in s: return 'AP'
    if s == 'Not Determined': return 'Not Determined'
    return s

merged['formula_norm'] = merged[formula_col].apply(normalise)
merged['our_norm']     = merged['Derived Dx'].fillna('')
mismatches = merged[merged['formula_norm'] != merged['our_norm']]
print(f'\n=== Verification vs existing Formula column ===')
print(f'Matching:   {len(merged) - len(mismatches)} / {len(merged)}')
print(f'Mismatches: {len(mismatches)}')
if len(mismatches) > 0:
    print(mismatches[['Project key', 'formula_norm', 'our_norm', 'Dx Source']].head(20).to_string())

# ── Formatting helpers ─────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='212529')
HDR_FONT   = Font(color='FFFFFF', bold=True)
PROXY_FILL = PatternFill('solid', fgColor='FFF2CC')   # yellow: proxy
REVIEW_FILL= PatternFill('solid', fgColor='FCE4D6')   # orange: needs review
NULL_FILL  = PatternFill('solid', fgColor='F2F2F2')   # grey: no dx
EVEN_FILL  = PatternFill('solid', fgColor='F8F9FA')
ODD_FILL   = PatternFill('solid', fgColor='FFFFFF')

DX_FILLS = {
    'PD':  PatternFill('solid', fgColor='DDEEFF'),
    'PSP': PatternFill('solid', fgColor='E2EFDA'),
    'MSA': PatternFill('solid', fgColor='E2EFDA'),
    'CBS': PatternFill('solid', fgColor='E2EFDA'),
    'DLB': PatternFill('solid', fgColor='E2EFDA'),
    'RBD': PatternFill('solid', fgColor='E2EFDA'),
    'ET':  PatternFill('solid', fgColor='E2EFDA'),
    'HC':  PatternFill('solid', fgColor='EAF4EA'),
    'AP':  PatternFill('solid', fgColor='F4EAFA'),
}

def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        dx_val    = str(row[2].value or '')   # Derived Dx col index 2
        proxy_val = row[4].value              # Is Proxy
        review_val= row[5].value              # Needs Review

        if review_val:
            fill = REVIEW_FILL
        elif proxy_val:
            fill = PROXY_FILL
        elif not dx_val:
            fill = NULL_FILL
        else:
            fill = DX_FILLS.get(dx_val, EVEN_FILL if i % 2 == 0 else ODD_FILL)

        for cell in row:
            cell.fill = fill

    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 60)

# ── Write output ───────────────────────────────────────────────────────────────
print(f'\nWriting {OUT_PATH}...')
with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    result.to_excel(writer,    sheet_name='PD_Diagnosis',  index=False)
    proxy_df.to_excel(writer,  sheet_name='Step2b_proxy',  index=False)
    review_df.to_excel(writer, sheet_name='Free_text_dx',  index=False)

wb = openpyxl.load_workbook(OUT_PATH)
style_sheet(wb['PD_Diagnosis'])

# Simple header-only styling for sub-sheets
for sn in ['Step2b_proxy', 'Free_text_dx']:
    ws = wb[sn]
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 60)

wb.save(OUT_PATH)
print('Done.')
print(f'\nSheets:')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row - 1} rows')
